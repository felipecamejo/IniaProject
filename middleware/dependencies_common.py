"""
Módulo común para gestión de dependencias.
Centraliza la instalación y verificación de paquetes requeridos.
"""


def asegurar_dependencias(paquetes: list, modulo_nombre: str = None):
    """
    Asegura que las dependencias estén instaladas.
    
    Args:
        paquetes: Lista de tuplas (nombre_paquete, nombre_instalacion)
                 Ej: [('sqlalchemy', 'SQLAlchemy'), ('pandas', 'pandas')]
        modulo_nombre: Nombre del módulo que requiere las dependencias (para mensajes)
    
    Returns:
        True si todas las dependencias están disponibles
    """
    try:
        from InstallDependencies import verificar_e_instalar, instalar_dependencias_faltantes
        INSTALL_DEPS_AVAILABLE = True
    except ImportError:
        INSTALL_DEPS_AVAILABLE = False
        print("InstallDependencies no disponible. Instala las dependencias manualmente.")
        return False
    
    faltantes = []
    for nombre_paquete, nombre_instalacion in paquetes:
        if not verificar_e_instalar(nombre_paquete, nombre_instalacion, silent=True):
            faltantes.append((nombre_paquete, nombre_instalacion))
    
    if faltantes:
        if modulo_nombre:
            print(f"Instalando dependencias faltantes para {modulo_nombre}...")
        for nombre_paquete, nombre_instalacion in faltantes:
            verificar_e_instalar(nombre_paquete, nombre_instalacion, silent=False)
    
    # Verificar importaciones
    for nombre_paquete, _ in paquetes:
        try:
            __import__(nombre_paquete)
        except ImportError:
            if INSTALL_DEPS_AVAILABLE:
                if instalar_dependencias_faltantes(modulo_nombre or 'módulo', silent=False):
                    try:
                        __import__(nombre_paquete)
                    except ImportError:
                        print(f"Error: No se pudo importar {nombre_paquete}. Instálalo manualmente.")
                        return False
                else:
                    return False
            else:
                print(f"Error: Falta el paquete '{nombre_paquete}'. Instálalo manualmente.")
                return False
    
    return True


def importar_sqlalchemy():
    """
    Importa SQLAlchemy asegurando que esté instalado.
    
    Returns:
        Tupla con (create_engine, text, inspect, Table, sessionmaker, automap_base)
    """
    asegurar_dependencias([('sqlalchemy', 'SQLAlchemy')], 'SQLAlchemy')
    
    from sqlalchemy import create_engine, text, inspect, Table
    from sqlalchemy.orm import sessionmaker
    from sqlalchemy.ext.automap import automap_base
    
    return create_engine, text, inspect, Table, sessionmaker, automap_base


def importar_pandas():
    """
    Importa pandas asegurando que esté instalado.
    
    Returns:
        Tupla con (pandas, numpy)
    """
    asegurar_dependencias([('pandas', 'pandas'), ('numpy', 'numpy')], 'Pandas')
    
    import pandas as pd
    import numpy as np
    
    return pd, np


def importar_openpyxl():
    """
    Importa openpyxl asegurando que esté instalado.
    
    Returns:
        Tupla con (success, Workbook, load_workbook, get_column_letter, Font, PatternFill, Border, Side, Alignment)
        Si success es False, todos los demás valores serán None
    """
    if asegurar_dependencias([('openpyxl', 'openpyxl')], 'OpenPyXL'):
        try:
            from openpyxl import Workbook, load_workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            return True, Workbook, load_workbook, get_column_letter, Font, PatternFill, Border, Side, Alignment
        except ImportError:
            return False, None, None, None, None, None, None, None, None
    return False, None, None, None, None, None, None, None, None

