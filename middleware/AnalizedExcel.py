import os
import argparse
import logging
from typing import Dict, List, Any, Optional, Tuple
from collections import defaultdict

# Intentar importar módulo de instalación de dependencias
try:
    from InstallDependencies import verificar_e_instalar, instalar_dependencias_faltantes
    INSTALL_DEPS_AVAILABLE = True
except ImportError:
    INSTALL_DEPS_AVAILABLE = False

# Verificar e instalar dependencias openpyxl
if INSTALL_DEPS_AVAILABLE:
    if not verificar_e_instalar('openpyxl', 'openpyxl', silent=True):
        print("Intentando instalar openpyxl...")
        verificar_e_instalar('openpyxl', 'openpyxl', silent=False)

# Dependencias para Excel
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False
    if INSTALL_DEPS_AVAILABLE:
        if verificar_e_instalar('openpyxl', 'openpyxl', silent=False):
            try:
                from openpyxl import load_workbook
                OPENPYXL_AVAILABLE = True
            except Exception:
                OPENPYXL_AVAILABLE = False

# Verificar e instalar dependencias pandas
if INSTALL_DEPS_AVAILABLE:
    if not verificar_e_instalar('pandas', 'pandas', silent=True):
        print("Intentando instalar pandas...")
        verificar_e_instalar('pandas', 'pandas', silent=False)

# Importaciones Pandas
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ModuleNotFoundError:
    PANDAS_AVAILABLE = False
    if INSTALL_DEPS_AVAILABLE:
        print("Instalando dependencias faltantes...")
        if instalar_dependencias_faltantes('AnalizedExcel', silent=False):
            import pandas as pd
            PANDAS_AVAILABLE = True
        else:
            print("No se pudieron instalar las dependencias. Instálalas manualmente con: pip install -r requirements.txt")
    else:
        print("Falta el paquete 'pandas'. Instálalo con: pip install pandas")

# Configuración de logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ================================
# MÓDULO: ANÁLISIS DE EXCEL
# ================================
def es_encabezado_generico(encabezado: str) -> bool:
    """
    Detecta si un encabezado es genérico (Columna_1, Columna_2, etc.).
    
    Args:
        encabezado: Nombre del encabezado a evaluar
        
    Returns:
        True si el encabezado es genérico, False en caso contrario
    """
    if not encabezado:
        return True
    
    encabezado_lower = encabezado.lower().strip()
    
    # Patrones de encabezados genéricos
    patrones_genericos = [
        r'^columna[_\s]?\d+$',
        r'^column[_\s]?\d+$',
        r'^col[_\s]?\d+$',
        r'^campo[_\s]?\d+$',
        r'^field[_\s]?\d+$',
        r'^f\d+$',
        r'^c\d+$'
    ]
    
    import re
    for patron in patrones_genericos:
        if re.match(patron, encabezado_lower):
            return True
    
    return False

def parece_encabezado_descriptivo(valor: Any) -> bool:
    """
    Evalúa si un valor parece ser un encabezado descriptivo.
    
    Args:
        valor: Valor a evaluar
        
    Returns:
        True si el valor parece ser un encabezado descriptivo
    """
    if valor is None:
        return False
    
    valor_str = str(valor).strip()
    
    # Si está vacío, no es descriptivo
    if not valor_str:
        return False
    
    # Si es solo un número, probablemente no es un encabezado
    try:
        float(valor_str)
        # Si es un número muy grande o muy pequeño, probablemente no es encabezado
        if abs(float(valor_str)) > 1000 or abs(float(valor_str)) < 0.0001:
            return False
    except (ValueError, TypeError):
        pass
    
    # Si tiene más de 2 palabras o contiene caracteres alfabéticos, probablemente es descriptivo
    palabras = valor_str.split()
    if len(palabras) >= 1 and any(c.isalpha() for c in valor_str):
        # Si tiene al menos 3 caracteres y no es solo un número
        if len(valor_str) >= 3:
            return True
    
    return False

def buscar_encabezados_alternativos(ws, max_filas_buscar: int = 5) -> Tuple[List[str], int]:
    """
    Busca encabezados descriptivos en las primeras filas del Excel.
    
    Args:
        ws: Worksheet de openpyxl
        max_filas_buscar: Número máximo de filas a buscar (por defecto 5)
        
    Returns:
        Tuple con (lista de encabezados encontrados, número de fila donde se encontraron)
        Si no se encuentran, retorna ([], 0)
    """
    encabezados_alternativos = []
    num_columnas = ws.max_column
    
    if num_columnas == 0:
        return encabezados_alternativos, 0
    
    # Buscar en las primeras filas (2-6, ya que la fila 1 ya se revisó)
    for fila in range(2, min(max_filas_buscar + 2, ws.max_row + 1)):
        fila_encabezados = []
        todos_descriptivos = True
        
        for col in range(1, num_columnas + 1):
            celda = ws.cell(row=fila, column=col)
            valor = celda.value
            
            if parece_encabezado_descriptivo(valor):
                fila_encabezados.append(str(valor).strip())
            else:
                todos_descriptivos = False
                break
        
        # Si encontramos una fila completa con encabezados descriptivos, usarla
        if todos_descriptivos and len(fila_encabezados) == num_columnas:
            logger.info(f"Encabezados descriptivos encontrados en la fila {fila}")
            return fila_encabezados, fila
    
    return encabezados_alternativos, 0

def obtener_encabezados_mejorados(ws) -> Tuple[List[str], int]:
    """
    Obtiene los encabezados del Excel, buscando en filas alternativas si la primera fila
    tiene encabezados genéricos.
    
    Args:
        ws: Worksheet de openpyxl
        
    Returns:
        Tuple con (lista de encabezados mejorados, número de fila donde están los encabezados)
        La fila será 1 si se usan los encabezados de la primera fila
    """
    encabezados = []
    num_columnas = ws.max_column
    
    if num_columnas == 0:
        return encabezados, 1
    
    # Obtener encabezados de la primera fila
    encabezados_fila1 = []
    for col in range(1, num_columnas + 1):
        celda = ws.cell(row=1, column=col)
        valor = celda.value
        if valor is not None:
            encabezados_fila1.append(str(valor).strip())
        else:
            encabezados_fila1.append(f"Columna_{col}")
    
    # Verificar si todos los encabezados son genéricos
    todos_genericos = all(es_encabezado_generico(enc) for enc in encabezados_fila1)
    
    if todos_genericos:
        logger.info("Encabezados genéricos detectados, buscando encabezados alternativos...")
        # Buscar encabezados alternativos en filas siguientes
        encabezados_alternativos, fila_encabezados = buscar_encabezados_alternativos(ws, max_filas_buscar=5)
        
        if encabezados_alternativos and len(encabezados_alternativos) == num_columnas:
            logger.info(f"Usando encabezados alternativos encontrados en la fila {fila_encabezados}")
            return encabezados_alternativos, fila_encabezados
        else:
            logger.info("No se encontraron encabezados alternativos descriptivos, usando encabezados genéricos")
            return encabezados_fila1, 1
    else:
        # Algunos encabezados son descriptivos, usar los de la primera fila
        # Pero reemplazar los genéricos con alternativas si es posible
        encabezados_mejorados = []
        encabezados_alternativos, fila_alternativos = buscar_encabezados_alternativos(ws, max_filas_buscar=5)
        
        for idx, enc in enumerate(encabezados_fila1):
            if es_encabezado_generico(enc) and encabezados_alternativos and idx < len(encabezados_alternativos):
                # Usar el encabezado alternativo si está disponible
                enc_alt = encabezados_alternativos[idx]
                if parece_encabezado_descriptivo(enc_alt):
                    encabezados_mejorados.append(enc_alt)
                else:
                    encabezados_mejorados.append(enc)
            else:
                encabezados_mejorados.append(enc)
        
        # Si se usaron alternativos, usar esa fila, sino usar fila 1
        fila_encabezados = fila_alternativos if fila_alternativos > 0 else 1
        return encabezados_mejorados, fila_encabezados

def analizar_estructura_excel(ruta_archivo: str) -> Dict[str, Any]:
    """
    Analiza la estructura de un archivo Excel y retorna información sobre sus hojas y columnas.
    
    Args:
        ruta_archivo: Ruta al archivo Excel a analizar
        
    Returns:
        Diccionario con información de la estructura del Excel
    """
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl no está instalado")
    
    if not os.path.exists(ruta_archivo):
        raise FileNotFoundError(f"El archivo no existe: {ruta_archivo}")
    
    estructura = {
        'archivo': os.path.basename(ruta_archivo),
        'ruta_completa': os.path.abspath(ruta_archivo),
        'hojas': []
    }
    
    try:
        wb = load_workbook(ruta_archivo, data_only=True)
        hojas = wb.sheetnames
        
        for nombre_hoja in hojas:
            ws = wb[nombre_hoja]
            info_hoja = {
                'nombre': nombre_hoja,
                'columnas': [],
                'total_filas': ws.max_row,
                'total_columnas': ws.max_column,
                'contexto': None
            }
            
            # Buscar títulos de sección para identificar el contexto (incluyendo nombre de hoja)
            contexto_seccion = buscar_titulos_seccion(ws, nombre_hoja=nombre_hoja, max_filas_buscar=10)
            if contexto_seccion:
                info_hoja['contexto'] = contexto_seccion
                logger.info(f"Contexto identificado para hoja '{nombre_hoja}': {contexto_seccion}")
            
            # Obtener encabezados mejorados (busca en filas alternativas si la primera fila tiene encabezados genéricos)
            encabezados = []
            fila_encabezados = 1
            if ws.max_row > 0 and ws.max_column > 0:
                encabezados, fila_encabezados = obtener_encabezados_mejorados(ws)
            
            # Calcular la fila de inicio de datos (después de los encabezados)
            fila_inicio_datos = fila_encabezados + 1
            
            # Analizar cada columna
            for idx, encabezado in enumerate(encabezados, start=1):
                columna_info = {
                    'nombre': encabezado,
                    'indice': idx,
                    'tipo_datos': [],
                    'valores_unicos': set(),
                    'valores_nulos': 0,
                    'total_valores': 0
                }
                
                # Analizar valores de la columna (desde la fila después de los encabezados)
                for row in range(fila_inicio_datos, ws.max_row + 1):
                    celda = ws.cell(row=row, column=idx)
                    valor = celda.value
                    columna_info['total_valores'] += 1
                    
                    # Ignorar errores de fórmula (como #DIV/0!, #VALUE!, etc.)
                    if isinstance(valor, str) and valor.startswith('#'):
                        # Es un error de fórmula, ignorarlo (contarlo como nulo)
                        columna_info['valores_nulos'] += 1
                        continue
                    
                    # Verificar si es un objeto de error de openpyxl
                    try:
                        from openpyxl.cell.cell import TYPE_ERROR
                        if celda.data_type == TYPE_ERROR:
                            # Es un error de fórmula, ignorarlo
                            columna_info['valores_nulos'] += 1
                            continue
                    except (ImportError, AttributeError):
                        pass
                    
                    if valor is None:
                        columna_info['valores_nulos'] += 1
                    else:
                        tipo_valor = type(valor).__name__
                        if tipo_valor not in columna_info['tipo_datos']:
                            columna_info['tipo_datos'].append(tipo_valor)
                        
                        # Agregar valor único (limitado a 100 para no sobrecargar)
                        if len(columna_info['valores_unicos']) < 100:
                            columna_info['valores_unicos'].add(str(valor))
                
                # Convertir set a lista para serialización
                columna_info['valores_unicos'] = sorted(list(columna_info['valores_unicos']))[:50]
                columna_info['tipo_datos'] = columna_info['tipo_datos']
                
                info_hoja['columnas'].append(columna_info)
            
            estructura['hojas'].append(info_hoja)
        
        wb.close()
        
    except Exception as e:
        logger.error(f"Error analizando Excel: {e}")
        raise
    
    return estructura

def buscar_titulos_seccion(ws, nombre_hoja: str = None, max_filas_buscar: int = 10) -> Optional[str]:
    """
    Busca títulos de sección en las primeras filas del Excel y en el nombre de la hoja
    para identificar el contexto de la entidad.
    
    Args:
        ws: Worksheet de openpyxl
        nombre_hoja: Nombre de la hoja del Excel
        max_filas_buscar: Número máximo de filas a buscar (por defecto 10)
        
    Returns:
        Nombre de la entidad identificada por el contexto, o None si no se encuentra
    """
    import re
    
    # Patrones de títulos de sección
    patrones_titulos = {
        'pureza': [
            r'pureza\s+ista',
            r'pureza\s+inase',
            r'cálculo\s+de\s+pureza',
            r'pureza\s+para',
            r'análisis\s+de\s+pureza',
            r'pureza\s+común',
            r'analisis\s+sp',  # "Analisis SP" (como en tu Excel)
            r'analisis\s+pureza'
        ],
        'pureza_pnotatum': [
            r'pureza\s+pnotatum',
            r'pureza\s+p\.?\s*notatum',
            r'pnotatum',
            r'p\.?\s*notatum',
            r'repeticiones\s+ppn'
        ],
        'dosn': [
            r'dosn',
            r'determinación\s+de\s+otras\s+semillas',
            r'análisis\s+dosn',
            r'determinacion\s+otras\s+semillas'
        ],
        'germinacion': [
            r'germinación',
            r'germinacion',
            r'análisis\s+de\s+germinación',
            r'prueba\s+de\s+germinación',
            r'analisis\s+germinacion'
        ],
        'tetrazolio': [
            r'tetrazolio',
            r'prueba\s+de\s+tetrazolio',
            r'análisis\s+de\s+viabilidad',
            r'viabilidad'
        ],
        'sanitario': [
            r'sanitario',
            r'análisis\s+sanitario',
            r'patógenos',
            r'hongos',
            r'analisis\s+sanitario'
        ],
        'pms': [
            r'pms',
            r'peso\s+de\s+mil\s+semillas',
            r'peso\s+1000\s+semillas',
            r'peso\s+mil\s+semillas'
        ]
    }
    
    # Buscar primero en el nombre de la hoja
    if nombre_hoja:
        nombre_hoja_lower = nombre_hoja.lower()
        for entidad, patrones in patrones_titulos.items():
            for patron in patrones:
                if re.search(patron, nombre_hoja_lower, re.IGNORECASE):
                    logger.info(f"Título de sección encontrado en nombre de hoja '{nombre_hoja}' -> Entidad: {entidad}")
                    return entidad
    
    # Buscar en las primeras filas
    for fila in range(1, min(max_filas_buscar + 1, ws.max_row + 1)):
        # Buscar en todas las columnas de la fila
        texto_fila = ""
        for col in range(1, min(ws.max_column + 1, 10)):  # Buscar en las primeras 10 columnas
            celda = ws.cell(row=fila, column=col)
            valor = celda.value
            if valor is not None:
                texto_fila += " " + str(valor).strip()
        
        texto_fila_lower = texto_fila.lower()
        
        # Verificar cada patrón
        for entidad, patrones in patrones_titulos.items():
            for patron in patrones:
                if re.search(patron, texto_fila_lower, re.IGNORECASE):
                    logger.info(f"Título de sección encontrado en fila {fila}: '{texto_fila.strip()}' -> Entidad: {entidad}")
                    return entidad
    
    return None

def mapear_columna_con_contexto(nombre_col: str, contexto: Optional[str]) -> Optional[str]:
    """
    Mapea el nombre de una columna a un nombre de columna de BD usando el contexto identificado.
    
    Args:
        nombre_col: Nombre de la columna del Excel
        contexto: Contexto identificado (pureza, pureza_pnotatum, dosn, etc.)
        
    Returns:
        Nombre de columna de BD sugerido, o None si no se puede mapear
    """
    nombre_col_lower = nombre_col.lower().strip()
    
    # Mapeos específicos por contexto
    mapeos_contexto = {
        'pureza': {
            # Mapeos de peso (peso de semilla usado en pureza)
            'peso': 'peso_inicial',
            'peso (g)': 'peso_inicial',
            'peso inicial': 'peso_inicial',
            'peso_inicial': 'peso_inicial',
            'peso inicial de la muestra': 'peso_inicial',
            'peso final': 'peso_inicial',  # Peso final de la muestra
            'peso final de la muestra': 'peso_inicial',
            'dif de peso': 'peso_inicial',  # Diferencia de peso
            'peso estimado': 'peso_inicial',  # Peso estimado de semillas
            'peso estimado (g)': 'peso_inicial',
            'peso del total de semillas analizadas': 'peso_inicial',
            'peso total de semillas contaminadas': 'peso_inicial',
            'control de pesos': 'peso_inicial',  # Control de pesos
            # Mapeos de semilla pura
            'semilla pura': 'semilla_pura',
            'semilla_pura': 'semilla_pura',
            'semillas puras': 'semilla_pura',
            'nº semillas puras': 'semilla_pura',
            'numero semillas puras': 'semilla_pura',
            'pu': 'semilla_pura',
            # Mapeos de materia inerte
            'materia inerte': 'material_inerte',
            'material inerte': 'material_inerte',
            'material_inerte': 'material_inerte',
            'materia_inerte': 'material_inerte',
            # Mapeos de otros cultivos
            'otros cultivos': 'otros_cultivos',
            'semilla cultivos': 'otros_cultivos',
            'semillas cultivos': 'otros_cultivos',
            'semilla cultivos': 'otros_cultivos',
            'otros_cultivos': 'otros_cultivos',
            # Mapeos de malezas
            'malezas': 'malezas',
            'semillas malezas': 'malezas',
            'semilla malezas': 'malezas',
            'semillas contaminadas': 'malezas',
            'semillas vanas': 'malezas',
            'hillas contaminadas y vanas': 'malezas',  # Nota: "hillas" parece ser "semillas"
            # Mapeos de semillas sanas (para pureza INASE)
            'semillas sanas': 'semilla_pura',  # Semillas sanas se mapean a semilla_pura
            'nº': 'semilla_pura',  # Número de semillas (contexto dependiente)
            # Mapeos de fechas
            'fecha inase': 'fecha_inase',
            'fecha_inase': 'fecha_inase',
            'fecha inia': 'fecha_inia',
            'fecha_inia': 'fecha_inia'
        },
        'pureza_pnotatum': {
            'repeticion': 'repeticion',
            'repeticiones': 'repeticion',
            'semillas puras': 'semillas_puras',
            'peso (g)': 'peso',
            'peso': 'peso',
            'semillas sanas': 'semillas_sanas',
            'contaminadas': 'semillas_contaminadas',
            'vanas': 'semillas_vanas'
        },
        'dosn': {
            'gramos analizados': 'gramos_analizados',
            'gramos_analizados': 'gramos_analizados',
            'tipos analisis': 'tipos_analisis',
            'determinacion': 'determinacion',
            'fecha inase': 'fecha_inase',
            'fecha inia': 'fecha_inia'
        },
        'germinacion': {
            'semillas germinadas': 'semillas_germinadas',
            'semillas no germinadas': 'semillas_no_germinadas',
            'porcentaje germinacion': 'porcentaje_germinacion',
            'fecha germinacion': 'fecha_germinacion'
        }
    }
    
    if contexto and contexto in mapeos_contexto:
        mapeos = mapeos_contexto[contexto]
        
        # Limpiar nombre de columna (remover caracteres especiales, espacios extra)
        import re
        nombre_col_limpio = re.sub(r'[^\w\s]', ' ', nombre_col_lower)
        nombre_col_limpio = re.sub(r'\s+', ' ', nombre_col_limpio).strip()
        
        # Buscar coincidencia exacta primero
        if nombre_col_lower in mapeos:
            return mapeos[nombre_col_lower]
        if nombre_col_limpio in mapeos:
            return mapeos[nombre_col_limpio]
        
        # Buscar coincidencia parcial (palabras clave)
        palabras_col = set(nombre_col_limpio.split())
        mejor_coincidencia = None
        mejor_puntuacion = 0
        
        for clave, valor in mapeos.items():
            palabras_clave = set(clave.split())
            # Calcular puntuación basada en palabras comunes
            palabras_comunes = palabras_col.intersection(palabras_clave)
            if palabras_comunes:
                puntuacion = len(palabras_comunes) / max(len(palabras_col), len(palabras_clave))
                if puntuacion > mejor_puntuacion and puntuacion >= 0.5:  # Al menos 50% de coincidencia
                    mejor_puntuacion = puntuacion
                    mejor_coincidencia = valor
        
        if mejor_coincidencia:
            return mejor_coincidencia
        
        # Buscar coincidencia parcial simple (substring)
        for clave, valor in mapeos.items():
            if clave in nombre_col_lower or nombre_col_lower in clave:
                return valor
    
    return None

def identificar_entidades_columnas(columnas: List[Dict[str, Any]], contexto: Optional[str] = None) -> Dict[str, List[str]]:
    """
    Identifica a qué entidades pertenecen las columnas basándose en patrones de nombres.
    
    Args:
        columnas: Lista de información de columnas
        
    Returns:
        Diccionario mapeando nombres de entidades a listas de columnas
    """
    # Patrones de identificación de entidades basados en el proyecto
    patrones_entidades = {
        'recibo': ['recibo', 'numero_muestra', 'numero_lote', 'fecha_ingreso', 'peso_kg', 'numero_envases'],
        'lote': ['lote', 'numero_lote', 'categoria', 'especie', 'cultivar'],
        'usuario': ['usuario', 'nombre', 'email', 'rol'],
        'dosn': ['dosn', 'gramos_analizados', 'tipos_analisis', 'determinacion', 'fecha_inia', 'fecha_inase'],
        'pureza': ['pureza', 'semilla_pura', 'material_inerte', 'otros_cultivos', 'malezas', 'fecha_inase', 'fecha_inia'],
        'pureza_pnotatum': ['pnotatum', 'pureza_pnotatum', 'repeticion'],
        'germinacion': ['germinacion', 'semillas_germinadas', 'semillas_no_germinadas', 'porcentaje_germinacion', 'fecha_germinacion'],
        'tetrazolio': ['tetrazolio', 'viables', 'no_viables', 'duras', 'concentracion', 'tincion'],
        'sanitario': ['sanitario', 'hongo', 'patogeno', 'enfermedad'],
        'pms': ['pms', 'gramos_pms', 'peso_muestra'],
        'deposito': ['deposito', 'nombre_deposito', 'ubicacion'],
        'certificado': ['certificado', 'numero_certificado', 'fecha_emision', 'firmante'],
        'maleza': ['maleza', 'nombre_maleza', 'tipo_maleza'],
        'cultivo': ['cultivo', 'nombre_cultivo'],
        'metodo': ['metodo', 'nombre_metodo', 'descripcion_metodo']
    }
    
    mapeo_entidades = defaultdict(list)
    
    # Si hay contexto, usarlo como entidad principal
    entidad_contexto = contexto
    
    for columna in columnas:
        nombre_col = columna['nombre'].lower()
        entidad_asignada = None
        mejor_coincidencia = 0
        
        # Si hay contexto, intentar mapear la columna usando el contexto
        if contexto:
            columna_mapeada = mapear_columna_con_contexto(nombre_col, contexto)
            if columna_mapeada:
                # Si se pudo mapear con el contexto, usar la entidad del contexto
                entidad_asignada = contexto
                logger.debug(f"Columna '{nombre_col}' mapeada a '{columna_mapeada}' usando contexto '{contexto}'")
        
        # Si no se asignó por contexto, buscar la mejor coincidencia con patrones
        if entidad_asignada is None:
            for entidad, patrones in patrones_entidades.items():
                coincidencias = sum(1 for patron in patrones if patron in nombre_col)
                if coincidencias > mejor_coincidencia:
                    mejor_coincidencia = coincidencias
                    entidad_asignada = entidad
        
        # Si no hay coincidencia clara, intentar identificar por prefijos comunes
        if entidad_asignada is None:
            if nombre_col.startswith('recibo_'):
                entidad_asignada = 'recibo'
            elif nombre_col.startswith('lote_'):
                entidad_asignada = 'lote'
            elif nombre_col.startswith('dosn_'):
                entidad_asignada = 'dosn'
            elif nombre_col.startswith('pureza_') or nombre_col.startswith('pnotatum_'):
                entidad_asignada = 'pureza_pnotatum' if 'pnotatum' in nombre_col else 'pureza'
            elif nombre_col.startswith('germinacion_'):
                entidad_asignada = 'germinacion'
            elif nombre_col.startswith('tetrazolio_'):
                entidad_asignada = 'tetrazolio'
            elif nombre_col.startswith('sanitario_'):
                entidad_asignada = 'sanitario'
            elif nombre_col.startswith('pms_'):
                entidad_asignada = 'pms'
            elif nombre_col.startswith('certificado_'):
                entidad_asignada = 'certificado'
            elif 'maleza' in nombre_col:
                entidad_asignada = 'maleza'
            elif 'cultivo' in nombre_col:
                entidad_asignada = 'cultivo'
            elif 'metodo' in nombre_col:
                entidad_asignada = 'metodo'
            elif 'usuario' in nombre_col or 'user' in nombre_col:
                entidad_asignada = 'usuario'
            elif 'deposito' in nombre_col:
                entidad_asignada = 'deposito'
            else:
                # Si hay contexto pero no se pudo mapear, usar el contexto como fallback
                if contexto:
                    entidad_asignada = contexto
                    logger.debug(f"Usando contexto '{contexto}' como fallback para columna '{nombre_col}'")
                else:
                    entidad_asignada = 'general'
        
        mapeo_entidades[entidad_asignada].append(columna['nombre'])
    
    # Si hay contexto y no se asignaron columnas, crear una entidad con el contexto
    if contexto and not mapeo_entidades:
        logger.info(f"Contexto identificado '{contexto}' pero no se asignaron columnas, creando entidad vacía")
        mapeo_entidades[contexto] = []
    
    return dict(mapeo_entidades)

def generar_mapeo_datos(estructura: Dict[str, Any]) -> Dict[str, Any]:
    """
    Genera un mapeo completo de datos del Excel, indicando qué datos tiene y a quiénes pertenecen.
    
    Args:
        estructura: Estructura del Excel analizada
        
    Returns:
        Mapeo completo de datos
    """
    mapeo_completo = {
        'archivo': estructura['archivo'],
        'ruta': estructura['ruta_completa'],
        'resumen': {
            'total_hojas': len(estructura['hojas']),
            'total_columnas': sum(len(hoja['columnas']) for hoja in estructura['hojas']),
            'total_filas': sum(hoja['total_filas'] for hoja in estructura['hojas'])
        },
        'hojas': []
    }
    
    for hoja_info in estructura['hojas']:
        hoja_mapeo = {
            'nombre': hoja_info['nombre'],
            'resumen': {
                'total_columnas': hoja_info['total_columnas'],
                'total_filas': hoja_info['total_filas']
            },
            'entidades': {}
        }
        
        # Identificar entidades de las columnas usando el contexto si está disponible
        contexto = hoja_info.get('contexto')
        entidades_columnas = identificar_entidades_columnas(hoja_info['columnas'], contexto=contexto)
        
        # Organizar datos por entidad
        for entidad, columnas_nombres in entidades_columnas.items():
            columnas_info = [col for col in hoja_info['columnas'] if col['nombre'] in columnas_nombres]
            
            entidad_info = {
                'nombre': entidad,
                'columnas': [],
                'total_columnas': len(columnas_info),
                'tipos_datos': set(),
                'valores_unicos_por_columna': {}
            }
            
            for col_info in columnas_info:
                col_mapeo = {
                    'nombre': col_info['nombre'],
                    'tipo_datos': col_info['tipo_datos'],
                    'valores_nulos': col_info['valores_nulos'],
                    'total_valores': col_info['total_valores'],
                    'porcentaje_nulos': round((col_info['valores_nulos'] / col_info['total_valores'] * 100) if col_info['total_valores'] > 0 else 0, 2),
                    'valores_unicos_muestra': col_info['valores_unicos'][:10]  # Solo primeros 10 para el mapeo
                }
                
                entidad_info['columnas'].append(col_mapeo)
                entidad_info['tipos_datos'].update(col_info['tipo_datos'])
                entidad_info['valores_unicos_por_columna'][col_info['nombre']] = len(col_info['valores_unicos'])
            
            entidad_info['tipos_datos'] = list(entidad_info['tipos_datos'])
            hoja_mapeo['entidades'][entidad] = entidad_info
        
        mapeo_completo['hojas'].append(hoja_mapeo)
    
    return mapeo_completo

def imprimir_mapeo(mapeo: Dict[str, Any], formato: str = 'texto') -> None:
    """
    Imprime el mapeo de datos en el formato especificado.
    
    Args:
        mapeo: Mapeo de datos a imprimir
        formato: Formato de salida ('texto' o 'json')
    """
    if formato == 'json':
        import json
        print(json.dumps(mapeo, indent=2, ensure_ascii=False))
    else:
        # Formato texto legible
        print("=" * 80)
        print(f"ANÁLISIS DE EXCEL: {mapeo['archivo']}")
        print("=" * 80)
        print(f"Ruta: {mapeo['ruta']}")
        print(f"\nResumen General:")
        print(f"  - Total de hojas: {mapeo['resumen']['total_hojas']}")
        print(f"  - Total de columnas: {mapeo['resumen']['total_columnas']}")
        print(f"  - Total de filas: {mapeo['resumen']['total_filas']}")
        print("\n" + "=" * 80)
        
        for hoja in mapeo['hojas']:
            print(f"\nHOJA: {hoja['nombre']}")
            print("-" * 80)
            print(f"  Columnas: {hoja['resumen']['total_columnas']}, Filas: {hoja['resumen']['total_filas']}")
            print("\n  Entidades identificadas:")
            
            for entidad_nombre, entidad_info in hoja['entidades'].items():
                print(f"\n    [{entidad_nombre.upper()}]")
                print(f"      Columnas: {entidad_info['total_columnas']}")
                print(f"      Tipos de datos: {', '.join(entidad_info['tipos_datos'])}")
                print(f"      Columnas detalladas:")
                
                for col in entidad_info['columnas']:
                    print(f"        - {col['nombre']}")
                    print(f"          Tipo: {', '.join(col['tipo_datos'])}")
                    print(f"          Valores nulos: {col['valores_nulos']} ({col['porcentaje_nulos']}%)")
                    if col['valores_unicos_muestra']:
                        muestra = ', '.join(col['valores_unicos_muestra'][:5])
                        print(f"          Muestra valores: {muestra}...")
        
        print("\n" + "=" * 80)

# ================================
# MÓDULO: FUNCIÓN PRINCIPAL
# ================================
def analizar_excel(ruta_archivo: str, formato_salida: str = 'texto') -> Dict[str, Any]:
    """
    Función principal que analiza un archivo Excel y genera el mapeo de datos.
    
    Args:
        ruta_archivo: Ruta al archivo Excel a analizar
        formato_salida: Formato de salida ('texto' o 'json')
        
    Returns:
        Mapeo completo de datos
    """
    logger.info(f"Iniciando análisis de Excel: {ruta_archivo}")
    
    try:
        # Analizar estructura
        estructura = analizar_estructura_excel(ruta_archivo)
        logger.info(f"Estructura analizada: {len(estructura['hojas'])} hojas")
        
        # Generar mapeo
        mapeo = generar_mapeo_datos(estructura)
        logger.info(f"Mapeo generado: {len(mapeo['hojas'])} hojas mapeadas")
        
        # Imprimir resultado
        imprimir_mapeo(mapeo, formato_salida)
        
        return mapeo
        
    except Exception as e:
        logger.error(f"Error analizando Excel: {e}")
        raise

def main():
    """Función principal del script."""
    parser = argparse.ArgumentParser(description="Analiza un archivo Excel y genera un mapeo de datos")
    parser.add_argument(
        "archivo",
        help="Ruta al archivo Excel a analizar"
    )
    parser.add_argument(
        "--formato",
        choices=["texto", "json"],
        default="texto",
        help="Formato de salida (texto por defecto)"
    )
    args = parser.parse_args()
    
    try:
        analizar_excel(args.archivo, args.formato)
    except Exception as e:
        logger.error(f"Error en el proceso principal: {e}")
        raise

if __name__ == "__main__":
    main()

