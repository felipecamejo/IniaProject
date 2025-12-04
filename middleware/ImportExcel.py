import os
import csv
import argparse
import logging
import re
import unicodedata
import traceback
from typing import Any, Dict, Iterable, List, Optional, Tuple
from datetime import datetime, date
from urllib.parse import quote_plus

# Importar configuración de metadatos
try:
    from metadata_config import (
        METADATA_KEYWORDS,
        METADATA_PATTERNS,
        VALID_HEADER_KEYWORDS,
        VALIDATION_CONFIG,
        ANALYSIS_TYPE_TO_TABLES,
        ANALYSIS_KEYWORDS,
        HEADER_SYNONYMS,
    )
except ImportError:
    # Fallback si no se puede importar
    METADATA_KEYWORDS = []
    METADATA_PATTERNS = []
    VALID_HEADER_KEYWORDS = []
    VALIDATION_CONFIG = {
        "min_headers_validos": 3,
        "max_header_length": 50,
        "min_header_length": 1,
        "max_numeric_percentage": 0.3,
        "max_metadata_percentage": 0.3,
        "min_valid_percentage": 0.5,
        "confidence_threshold": 0.4,
        "max_title_length": 30,
    }
    ANALYSIS_TYPE_TO_TABLES = {}
    ANALYSIS_KEYWORDS = {}
    HEADER_SYNONYMS = {}

# Intentar importar módulo de instalación de dependencias
try:
    from InstallDependencies import verificar_e_instalar, instalar_dependencias_faltantes
    INSTALL_DEPS_AVAILABLE = True
except ImportError:
    INSTALL_DEPS_AVAILABLE = False

# Verificar e instalar dependencias SQLAlchemy
if INSTALL_DEPS_AVAILABLE:
    if not verificar_e_instalar('sqlalchemy', 'SQLAlchemy', silent=True):
        print("Intentando instalar SQLAlchemy...")
        verificar_e_instalar('sqlalchemy', 'SQLAlchemy', silent=False)

# Importaciones SQLAlchemy
try:
    from sqlalchemy import create_engine, text, Table, inspect
    from sqlalchemy.orm import sessionmaker
    from sqlalchemy.ext.automap import automap_base
except ModuleNotFoundError:
    if INSTALL_DEPS_AVAILABLE:
        print("Instalando dependencias faltantes...")
        if instalar_dependencias_faltantes('ImportExcel', silent=False):
            from sqlalchemy import create_engine, text, Table, inspect
            from sqlalchemy.orm import sessionmaker
            from sqlalchemy.ext.automap import automap_base
        else:
            print("No se pudieron instalar las dependencias. Instálalas manualmente con: pip install -r requirements.txt")
            raise
    else:
        print("Falta el paquete 'sqlalchemy'. Instálalo con: pip install SQLAlchemy")
        raise

# Dependencias opcionales para Excel
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False
    # Intentar instalar openpyxl si el módulo de instalación está disponible
    if INSTALL_DEPS_AVAILABLE:
        if verificar_e_instalar('openpyxl', 'openpyxl', silent=False):
            try:
                from openpyxl import load_workbook
                OPENPYXL_AVAILABLE = True
            except Exception:
                OPENPYXL_AVAILABLE = False

# Configuración de logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Importar configuración centralizada de base de datos
from database_config import build_connection_string

def obtener_engine():
    """Obtiene un engine de SQLAlchemy configurado."""
    try:
        from app.services.database_service import create_engine_with_pool
        return create_engine_with_pool()
    except ImportError:
        # Fallback si no está disponible el servicio
        connection_string = build_connection_string()
        return create_engine(connection_string)

# ================================
# MÓDULO: AUTOMAPEO DE MODELOS
# ================================
Base = None
_engine = None
_models_initialized = False
MODELS = {}
# Caché de tablas para evitar acceder al mapper repetidamente
_TABLES_CACHE = {}

def _intentar_reflejar_tablas_selectivas(engine, tablas_criticas=None):
    """
    Intenta reflejar tablas de forma selectiva para evitar problemas de codificación.
    
    Args:
        engine: Engine de SQLAlchemy
        tablas_criticas: Lista de nombres de tablas críticas a reflejar primero
    
    Returns:
        Base preparado o None si falla
    """
    if tablas_criticas is None:
        tablas_criticas = ['lote', 'recibo']  # Tablas críticas para importación
    
    try:
        from sqlalchemy import inspect as sql_inspect
        inspector = sql_inspect(engine)
        
        # Obtener esquemas disponibles
        schemas = inspector.get_schema_names()
        tablas_disponibles = []
        
        for schema in schemas:
            if schema in ['pg_catalog', 'information_schema']:
                continue
            try:
                tablas = inspector.get_table_names(schema=schema)
                for tabla in tablas:
                    tablas_disponibles.append((schema, tabla))
            except Exception as e:
                logger.warning(f"Error obteniendo tablas del esquema {schema}: {e}")
                continue
        
        # Intentar reflejar solo tablas críticas primero
        tablas_a_reflejar = []
        for schema, tabla in tablas_disponibles:
            tabla_lower = tabla.lower()
            if any(critica.lower() in tabla_lower for critica in tablas_criticas):
                tablas_a_reflejar.append((schema, tabla))
        
        if tablas_a_reflejar:
            logger.info(f"Intentando reflejar {len(tablas_a_reflejar)} tablas críticas...")
            Base = automap_base()
            metadata = Base.metadata
            
            # Reflejar solo las tablas críticas
            nombres_tablas = [tabla for _, tabla in tablas_a_reflejar]
            metadata.reflect(engine, only=nombres_tablas)
            Base.prepare()
            logger.info(f"✓ Tablas críticas reflejadas exitosamente: {len(Base.classes)} tablas")
            return Base
        
        # Si no hay tablas críticas, intentar reflejar todas
        logger.info("No se encontraron tablas críticas, intentando reflejar todas...")
        Base = automap_base()
        Base.prepare(autoload_with=engine, generate_relationship=None)
        return Base
        
    except Exception as e:
        logger.error(f"Error en reflejo selectivo: {e}")
        return None


def inicializar_automap(engine=None):
    """Inicializa automap_base y genera modelos automáticamente desde la BD."""
    global Base, _engine, _models_initialized, MODELS
    
    if _models_initialized and Base is not None:
        return Base
    
    if engine is None:
        # Usar función centralizada para crear engine con codificación UTF-8
        try:
            from app.services.database_service import create_engine_with_pool
            _engine = create_engine_with_pool()
            logger.info("Engine creado usando create_engine_with_pool() con codificación UTF-8")
        except Exception as e:
            logger.error(f"Error creando engine con create_engine_with_pool(): {e}")
            logger.error(traceback.format_exc())
            raise
    else:
        _engine = engine
    
    Base = automap_base()
    
    logger.info("Iniciando preparación de automap_base...")
    logger.debug(f"Engine configurado: {_engine.url if hasattr(_engine, 'url') else 'N/A'}")
    
    try:
        # Deshabilitar generación automática de relaciones para evitar conflictos de backref
        # Solo necesitamos las columnas para la importación, no las relaciones
        logger.info("Ejecutando Base.prepare() para reflejar todas las tablas...")
        Base.prepare(
            autoload_with=_engine,
            generate_relationship=None  # No generar relaciones automáticamente
            # Nota: reflect=True está deprecado cuando se usa autoload_with (reflexión automática)
        )
        logger.info(f"Modelos generados automáticamente: {len(Base.classes)} tablas")
    except (UnicodeDecodeError, UnicodeError) as e:
        logger.error("=" * 80)
        logger.error("ERROR DE CODIFICACIÓN DETECTADO DURANTE Base.prepare()")
        logger.error("=" * 80)
        logger.error(f"Tipo de error: {type(e).__name__}")
        logger.error(f"Mensaje: {e}")
        if hasattr(e, 'start') and hasattr(e, 'end'):
            logger.error(f"Posición del error: bytes {e.start}-{e.end}")
        if hasattr(e, 'object') and hasattr(e, 'start'):
            try:
                byte_problematico = e.object[e.start] if e.start < len(e.object) else None
                logger.error(f"Byte problemático: {hex(byte_problematico) if byte_problematico is not None else 'N/A'}")
                # Intentar mostrar contexto alrededor del byte problemático
                inicio = max(0, e.start - 10)
                fin = min(len(e.object), e.end + 10)
                contexto = e.object[inicio:fin]
                logger.error(f"Contexto (bytes): {contexto}")
            except Exception:
                pass
        logger.error("Stack trace completo:")
        logger.error(traceback.format_exc())
        logger.warning("Esto puede deberse a nombres de tablas/columnas con caracteres especiales.")
        logger.warning("Intentando continuar con manejo de errores de codificación...")
        # Verificar si se cargaron algunas clases a pesar del error
        if not hasattr(Base, 'classes') or len(Base.classes) == 0:
            # Si no se cargaron clases, intentar con configuración de codificación diferente
            logger.warning("No se cargaron clases. Intentando con configuración alternativa...")
            try:
                # Intentar con encoding explícito en la conexión
                Base = automap_base()
                logger.info("Reintentando Base.prepare() con configuración alternativa...")
                Base.prepare(autoload_with=_engine, generate_relationship=None)
                logger.info(f"Modelos generados automáticamente (con encoding alternativo): {len(Base.classes)} tablas")
            except Exception as e2:
                logger.error(f"Error en intento alternativo: {e2}")
                logger.error("Stack trace del intento alternativo:")
                logger.error(traceback.format_exc())
                
                # Intentar reflejo selectivo como último recurso
                logger.warning("Intentando reflejo selectivo de tablas críticas...")
                Base_selectivo = _intentar_reflejar_tablas_selectivas(_engine)
                if Base_selectivo and hasattr(Base_selectivo, 'classes') and len(Base_selectivo.classes) > 0:
                    Base = Base_selectivo
                    logger.info(f"✓ Reflejo selectivo exitoso: {len(Base.classes)} tablas cargadas")
                else:
                    raise RuntimeError(f"No se pudieron cargar modelos debido a error de codificación: {e}. Error alternativo: {e2}")
    except Exception as e:
        logger.error("=" * 80)
        logger.error("ERROR INESPERADO DURANTE Base.prepare()")
        logger.error("=" * 80)
        logger.error(f"Tipo de error: {type(e).__name__}")
        logger.error(f"Mensaje: {e}")
        logger.error("Stack trace completo:")
        logger.error(traceback.format_exc())
        # Si falla, intentar sin reflect=True como fallback
        try:
            logger.warning("Intentando inicializar automap sin reflect=True...")
            Base = automap_base()
            Base.prepare(autoload_with=_engine, generate_relationship=None)
            logger.info(f"Modelos generados automáticamente (fallback): {len(Base.classes)} tablas")
        except (UnicodeDecodeError, UnicodeError) as e2:
            logger.error(f"Error de codificación en fallback de automap: {e2}")
            logger.error("Stack trace del fallback:")
            logger.error(traceback.format_exc())
            logger.warning("Continuando con manejo de errores de codificación...")
            pass
        except Exception as e2:
            logger.error(f"Error en fallback de automap: {e2}")
            logger.error("Stack trace del fallback:")
            logger.error(traceback.format_exc())
            
            # Intentar reflejo selectivo como último recurso
            logger.warning("Intentando reflejo selectivo de tablas críticas como último recurso...")
            Base_selectivo = _intentar_reflejar_tablas_selectivas(_engine)
            if Base_selectivo and hasattr(Base_selectivo, 'classes') and len(Base_selectivo.classes) > 0:
                Base = Base_selectivo
                logger.info(f"✓ Reflejo selectivo exitoso: {len(Base.classes)} tablas cargadas")
            else:
                raise
    
    # Verificar que se cargaron clases
    if not hasattr(Base, 'classes') or len(Base.classes) == 0:
        raise RuntimeError("No se cargaron modelos de la base de datos. Base.classes está vacío.")
    
    MODELS.clear()
    modelos_procesados = 0
    modelos_con_error = 0
    
    for class_name in dir(Base.classes):
        if not class_name.startswith('_'):
            try:
                cls = getattr(Base.classes, class_name)
                # Intentar obtener el nombre de la tabla de diferentes formas
                tabla_nombre = None
                
                # Método 1: Intentar __tablename__
                try:
                    if hasattr(cls, '__tablename__'):
                        tabla_nombre_raw = cls.__tablename__
                        # Manejar posibles problemas de codificación
                        if isinstance(tabla_nombre_raw, bytes):
                            tabla_nombre = tabla_nombre_raw.decode('utf-8', errors='replace').lower()
                        else:
                            tabla_nombre = str(tabla_nombre_raw).lower()
                except (UnicodeDecodeError, UnicodeError) as e:
                    logger.warning(f"Error de codificación obteniendo __tablename__ de {class_name}: {e}")
                    try:
                        # Intentar con diferentes codificaciones
                        if isinstance(cls.__tablename__, bytes):
                            tabla_nombre = cls.__tablename__.decode('latin-1', errors='replace').lower()
                    except:
                        pass
                except Exception:
                    pass
                
                # Método 2: Intentar __table__.name
                if tabla_nombre is None:
                    try:
                        if hasattr(cls, '__table__') and hasattr(cls.__table__, 'name'):
                            tabla_nombre_raw = cls.__table__.name
                            # Manejar posibles problemas de codificación
                            if isinstance(tabla_nombre_raw, bytes):
                                tabla_nombre = tabla_nombre_raw.decode('utf-8', errors='replace').lower()
                            else:
                                tabla_nombre = str(tabla_nombre_raw).lower()
                    except (UnicodeDecodeError, UnicodeError) as e:
                        logger.warning(f"Error de codificación obteniendo __table__.name de {class_name}: {e}")
                        try:
                            if hasattr(cls, '__table__') and hasattr(cls.__table__, 'name'):
                                if isinstance(cls.__table__.name, bytes):
                                    tabla_nombre = cls.__table__.name.decode('latin-1', errors='replace').lower()
                        except:
                            pass
                    except Exception:
                        pass
                
                # Método 3: Usar el nombre de la clase directamente (en automap suelen coincidir)
                if tabla_nombre is None:
                    try:
                        tabla_nombre = str(class_name).lower()
                    except Exception:
                        logger.warning(f"Error convirtiendo class_name a string: {class_name}")
                        continue
                
                # Mapear la clase
                if tabla_nombre:
                    MODELS[tabla_nombre] = cls
                    MODELS[class_name.lower()] = cls
                    modelos_procesados += 1
                    logger.debug(f"Mapeado: {tabla_nombre} -> {class_name}")
            except (UnicodeDecodeError, UnicodeError) as e:
                modelos_con_error += 1
                logger.warning(f"Error de codificación procesando clase {class_name}: {e}. Omitiendo esta clase.")
                continue
            except Exception as e:
                modelos_con_error += 1
                logger.warning(f"Error procesando clase {class_name}: {e}")
                continue
    
    logger.info(f"Total de modelos mapeados: {len(MODELS)} (procesados: {modelos_procesados}, con errores: {modelos_con_error})")
    
    # Verificar que se mapearon al menos algunos modelos
    if len(MODELS) == 0:
        raise RuntimeError("No se pudieron mapear modelos. MODELS está vacío después del procesamiento.")
    _models_initialized = True
    return Base

def obtener_modelo(nombre_tabla):
    """Obtiene un modelo por nombre de tabla."""
    if not _models_initialized or Base is None:
        inicializar_automap()
    
    nombre_tabla_lower = nombre_tabla.lower()
    if nombre_tabla_lower in MODELS:
        return MODELS[nombre_tabla_lower]
    
    if Base is not None:
        for class_name in dir(Base.classes):
            if not class_name.startswith('_'):
                try:
                    cls = getattr(Base.classes, class_name)
                    if hasattr(cls, '__tablename__') and cls.__tablename__.lower() == nombre_tabla_lower:
                        MODELS[nombre_tabla_lower] = cls
                        return cls
                except Exception:
                    continue
    
    raise AttributeError(f"Tabla '{nombre_tabla}' no encontrada en modelos reflejados")

def obtener_nombre_tabla_seguro(model) -> str:
    """
    Obtiene el nombre de la tabla de un modelo de forma segura.
    Maneja diferentes formas en que SQLAlchemy puede exponer el nombre de la tabla.
    """
    try:
        # Método 1: Intentar __tablename__ directamente
        if hasattr(model, '__tablename__'):
            return model.__tablename__
    except (AttributeError, TypeError):
        pass
    
    try:
        # Método 2: Intentar __table__.name
        if hasattr(model, '__table__') and hasattr(model.__table__, 'name'):
            return model.__table__.name
    except (AttributeError, TypeError):
        pass
    
    try:
        # Método 3: Usar el nombre de la clase como fallback
        if hasattr(model, '__name__'):
            return model.__name__.lower()
    except (AttributeError, TypeError):
        pass
    
    try:
        # Método 4: Intentar obtener desde el mapeo
        if hasattr(model, '__mapper__') and hasattr(model.__mapper__, 'tables'):
            tables = model.__mapper__.tables
            if tables:
                return list(tables)[0].name
    except (AttributeError, TypeError, IndexError):
        pass
    
    # Último recurso: usar el nombre de la clase como string
    return str(model).split('.')[-1].split("'")[0].lower()

# ================================
# MÓDULO: UTILIDADES DE ARCHIVOS
# ================================
def detect_format_from_path(ruta_archivo: str) -> Optional[str]:
    """Detecta el formato de un archivo basándose en su extensión."""
    if not ruta_archivo:
        return None
    
    _, ext = os.path.splitext(ruta_archivo.lower())
    if ext == '.xlsx' or ext == '.xls':
        return 'xlsx'
    elif ext == '.csv':
        return 'csv'
    return None

def read_rows_from_xlsx(ruta_archivo: str, max_rows: Optional[int] = None) -> Tuple[List[str], List[List[Any]]]:
    """Lee las filas de un archivo Excel. Retorna (headers, rows)."""
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl no está instalado")
    
    headers = []
    rows = []
    
    try:
        wb = load_workbook(ruta_archivo, data_only=True)
        ws = wb.active
        
        # Validar que el archivo tiene contenido
        if ws.max_row == 0 or ws.max_column == 0:
            wb.close()
            raise ValueError("El archivo Excel está vacío o no tiene datos")
        
        # Leer encabezados (primera fila)
        if ws.max_row > 0:
            for col in range(1, ws.max_column + 1):
                celda = ws.cell(row=1, column=col)
                valor = celda.value
                headers.append(str(valor).strip() if valor is not None else f"Columna_{col}")
        
        # Validar que hay headers válidos
        headers_validos = [h for h in headers if h and h.strip() and not h.startswith("Columna_")]
        if len(headers_validos) < 3:
            wb.close()
            raise ValueError(f"El archivo Excel no tiene suficientes encabezados válidos. Encontrados: {len(headers_validos)}, mínimo requerido: 3")
        
        # Leer filas de datos
        start_row = 2
        end_row = ws.max_row + 1
        if max_rows:
            end_row = min(start_row + max_rows, end_row)
        
        for row_num in range(start_row, end_row):
            fila = []
            for col in range(1, ws.max_column + 1):
                celda = ws.cell(row=row_num, column=col)
                fila.append(celda.value)
            rows.append(fila)
        
        wb.close()
        
        # Validar que hay datos (solo si no es max_rows limitado)
        if not max_rows:
            filas_con_datos = [f for f in rows if any(v is not None and str(v).strip() for v in f)]
            if not filas_con_datos:
                logger.warning("El archivo Excel tiene encabezados pero no tiene filas con datos")
    except ValueError:
        raise
    except Exception as e:
        logger.error(f"Error leyendo archivo Excel: {e}")
        raise
    
    return headers, rows

def read_rows_from_csv(ruta_archivo: str, max_rows: Optional[int] = None) -> Tuple[List[str], List[List[Any]]]:
    """Lee las filas de un archivo CSV. Retorna (headers, rows)."""
    headers = []
    rows = []
    
    try:
        with open(ruta_archivo, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            
            # Leer encabezados (primera fila)
            try:
                headers = next(reader)
                headers = [h.strip() if h else f"Columna_{i+1}" for i, h in enumerate(headers)]
            except StopIteration:
                raise ValueError("El archivo CSV está vacío")
            
            # Validar que hay headers válidos
            headers_validos = [h for h in headers if h and h.strip() and not h.startswith("Columna_")]
            if len(headers_validos) < 3:
                raise ValueError(f"El archivo CSV no tiene suficientes encabezados válidos. Encontrados: {len(headers_validos)}, mínimo requerido: 3")
            
            # Leer filas de datos
            for i, fila in enumerate(reader):
                if max_rows and i >= max_rows:
                    break
                rows.append(fila)
            
            # Validar que hay datos
            filas_con_datos = [f for f in rows if any(v is not None and str(v).strip() for v in f)]
            if not filas_con_datos:
                logger.warning("El archivo CSV tiene encabezados pero no tiene filas con datos")
    except ValueError:
        raise
    except Exception as e:
        logger.error(f"Error leyendo archivo CSV: {e}")
        raise
    
    return headers, rows

def normalize_header_names(headers: List[str]) -> List[str]:
    """Normaliza los nombres de los encabezados."""
    normalized = []
    for header in headers:
        if header:
            # Normalizar unicode
            header = unicodedata.normalize('NFKD', str(header))
            # Convertir a minúsculas y reemplazar espacios/guiones con guiones bajos
            header = header.lower().strip()
            header = re.sub(r'[\s\-]+', '_', header)
            # Remover caracteres especiales
            header = re.sub(r'[^a-z0-9_]', '', header)
            # Remover guiones bajos múltiples
            header = re.sub(r'_+', '_', header)
            # Remover guiones bajos al inicio y final
            header = header.strip('_')
            if not header:
                header = f"columna_{len(normalized)+1}"
        else:
            header = f"columna_{len(normalized)+1}"
        normalized.append(header)
    return normalized

# ================================
# MÓDULO: DETECCIÓN DE TABLAS
# ================================

# Mapeo de nombres de archivos a tablas (para mejorar detección automática)
MAPEO_NOMBRES_ARCHIVOS = {
    # Tablas principales (coinciden con nombre de archivo)
    'autocompletado': 'autocompletado',
    'certificado': 'certificado',
    'conteo_germinacion': 'conteo_germinacion',
    'cultivo': 'cultivo',
    'deposito': 'deposito',
    'dosn': 'dosn',
    'dosn_cultivo': 'dosn_cultivo',
    'dosn_maleza': 'dosn_maleza',
    'germinacion': 'germinacion',
    'germinacion_curada_laboratorio': 'germinacion_curada_laboratorio',
    'germinacion_curada_planta': 'germinacion_curada_planta',
    'germinacion_normal_por_conteo': 'germinacion_normal_por_conteo',
    'germinacion_sin_curar': 'germinacion_sin_curar',
    'gramos_pms': 'gramos_pms',
    'hongo': 'hongo',
    'humedad_recibo': 'humedad_recibo',
    'lote': 'lote',
    'maleza': 'maleza',
    'metodo': 'metodo',
    'pms': 'pms',
    'pureza': 'pureza',
    'pureza_pnotatum': 'pureza_pnotatum',
    'recibo': 'recibo',
    'repeticiones_pureza_pnotatum': 'repeticiones_pureza_pnotatum',
    'sanitario': 'sanitario',
    'sanitario_hongo': 'sanitario_hongo',
    'tetrazolio': 'tetrazolio',
    'tetrazolio_detalle_semillas': 'tetrazolio_detalle_semillas',
    'usuario': 'usuario',
    'usuario_lote': 'usuario_lote',
    'viabilidad_reps_tetrazolio': 'viabilidad_reps_tetrazolio',
    'logs': 'logs',
    
    # Variantes de dosn_cultivo (todos van a dosn_cultivo)
    'dosn_cultivo_inase': 'dosn_cultivo',
    'dosn_cultivo_inia': 'dosn_cultivo',
    
    # Variantes de dosn_maleza (todos van a dosn_maleza)
    'dosn_maleza_normal_inase': 'dosn_maleza',
    'dosn_maleza_normal_inia': 'dosn_maleza',
    'dosn_maleza_tolerada_inase': 'dosn_maleza',
    'dosn_maleza_tolerada_inia': 'dosn_maleza',
    'dosn_maleza_tolerancia_cero_inase': 'dosn_maleza',
    'dosn_maleza_tolerancia_cero_inia': 'dosn_maleza',
    
    # Variantes de pureza_maleza (todos van a pureza_maleza)
    'pureza_maleza_normal': 'pureza_maleza',
    'pureza_maleza_tolerada': 'pureza_maleza',
    'pureza_maleza_tolerancia_cero': 'pureza_maleza',
}

def detectar_tabla_por_nombre_archivo(nombre_archivo: str) -> Optional[str]:
    """
    Detecta la tabla destino basándose en el nombre del archivo.
    
    Args:
        nombre_archivo: Nombre del archivo (con o sin extensión)
    
    Returns:
        Nombre de la tabla detectada o None si no se encuentra
    """
    if not nombre_archivo:
        return None
    
    # Extraer nombre sin extensión
    nombre_sin_ext = os.path.splitext(nombre_archivo)[0].lower().strip()
    
    # Buscar en mapeo directo
    if nombre_sin_ext in MAPEO_NOMBRES_ARCHIVOS:
        tabla_mapeada = MAPEO_NOMBRES_ARCHIVOS[nombre_sin_ext]
        logger.debug(f"Tabla detectada por mapeo de nombre: '{nombre_sin_ext}' -> '{tabla_mapeada}'")
        return tabla_mapeada
    
    # Buscar coincidencias parciales (para casos como "dosn_cultivo_inase.xlsx")
    for nombre_archivo_mapeo, tabla in MAPEO_NOMBRES_ARCHIVOS.items():
        if nombre_archivo_mapeo in nombre_sin_ext or nombre_sin_ext in nombre_archivo_mapeo:
            logger.debug(f"Tabla detectada por coincidencia parcial: '{nombre_sin_ext}' -> '{tabla}' (mapeo: '{nombre_archivo_mapeo}')")
            return tabla
    
    # Si no hay mapeo, verificar si el nombre coincide directamente con una tabla en MODELS
    if nombre_sin_ext in MODELS:
        logger.debug(f"Tabla detectada por nombre directo en MODELS: '{nombre_sin_ext}'")
        return nombre_sin_ext
    
    return None

def detectar_tipo_analisis_por_contenido(ruta_archivo: str) -> Optional[str]:
    """Detecta el tipo de análisis basándose en el contenido del archivo."""
    try:
        fmt = detect_format_from_path(ruta_archivo)
        if not fmt:
            return None
        
        # Leer primera fila para detectar tipo
        if fmt == 'xlsx':
            headers, _ = read_rows_from_xlsx(ruta_archivo, max_rows=1)
        else:
            headers, _ = read_rows_from_csv(ruta_archivo, max_rows=1)
        
        headers_lower = [h.lower() for h in headers if h]
        
        # Detectar tipo de análisis basándose en keywords
        if any('dosn' in h for h in headers_lower):
            return 'dosn'
        elif any('pureza' in h for h in headers_lower):
            return 'pureza'
        elif any('germinacion' in h for h in headers_lower):
            return 'germinacion'
        elif any('tetrazolio' in h for h in headers_lower):
            return 'tetrazolio'
        elif any('sanitario' in h for h in headers_lower):
            return 'sanitario'
        elif any('pms' in h for h in headers_lower):
            return 'pms'
        
        return None
    except Exception as e:
        logger.warning(f"Error detectando tipo de análisis: {e}")
        return None

def detectar_tabla_por_columnas(session, headers: List[str], tipo_analisis: Optional[str] = None) -> Optional[str]:
    """Detecta la tabla basándose en las columnas del archivo."""
    if not headers or len(headers) < 3:
        return None
    
    headers_lower = [h.lower() for h in headers if h]
    
    # Buscar en los modelos disponibles
    mejor_coincidencia = None
    mejor_puntuacion = 0
    
    for tabla_nombre, model in MODELS.items():
        try:
            # Obtener columnas del modelo
            columnas_modelo = [c.name.lower() for c in model.__table__.columns]
            
            # Calcular coincidencias
            coincidencias = sum(1 for h in headers_lower if h in columnas_modelo)
            puntuacion = coincidencias / len(headers_lower) if headers_lower else 0
            
            if puntuacion > mejor_puntuacion and puntuacion >= 0.3:  # Al menos 30% de coincidencia
                mejor_puntuacion = puntuacion
                mejor_coincidencia = tabla_nombre
        except Exception as e:
            logger.debug(f"Error verificando tabla {tabla_nombre}: {e}")
            continue
    
    return mejor_coincidencia

# ================================
# MÓDULO: SINCRONIZACIÓN DE SECUENCIAS
# ================================
def asegurar_autoincrementos(engine):
    """Asegura que las secuencias de autoincremento estén sincronizadas."""
    try:
        with engine.connect() as conn:
            # Obtener todas las tablas con columnas serial/bigserial
            query = text("""
                SELECT 
                    t.table_name,
                    c.column_name,
                    c.data_type
                FROM information_schema.tables t
                JOIN information_schema.columns c ON t.table_name = c.table_name
                WHERE t.table_schema = 'public'
                    AND t.table_type = 'BASE TABLE'
                    AND c.data_type IN ('integer', 'bigint')
                    AND c.column_default LIKE 'nextval%'
                ORDER BY t.table_name, c.column_name
            """)
            
            result = conn.execute(query)
            for row in result:
                tabla = row[0]
                columna = row[1]
                
                # Obtener el valor máximo actual
                query_max = text(f"SELECT COALESCE(MAX({columna}), 0) FROM {tabla}")
                max_val = conn.execute(query_max).scalar()
                
                # Obtener el nombre de la secuencia
                query_seq = text(f"""
                    SELECT pg_get_serial_sequence('{tabla}', '{columna}')
                """)
                seq_name = conn.execute(query_seq).scalar()
                
                if seq_name:
                    # Sincronizar la secuencia
                    query_sync = text(f"SELECT setval('{seq_name}', {max_val}, true)")
                    conn.execute(query_sync)
                    conn.commit()
                    logger.debug(f"Secuencia {seq_name} sincronizada a {max_val}")
    except Exception as e:
        logger.warning(f"Error sincronizando secuencias: {e}")

# ================================
# MÓDULO: IMPORTACIÓN DE ARCHIVOS
# ================================
def validar_foreign_keys(session, tabla_nombre: str, datos: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Valida que los valores de foreign keys existen en las tablas referenciadas.
    
    Args:
        session: Sesión de SQLAlchemy
        tabla_nombre: Nombre de la tabla
        datos: Diccionario con los datos a validar
    
    Returns:
        List[Dict[str, Any]]: Lista de errores encontrados. Cada error tiene:
            - 'columna': nombre de la columna FK
            - 'valor': valor que no existe
            - 'tabla_referenciada': tabla que debería contener el valor
            - 'mensaje': mensaje de error descriptivo
    """
    errores = []
    
    try:
        from sqlalchemy import text, MetaData, inspect
        
        # Obtener constraints de foreign key de la tabla
        query_fk = text("""
            SELECT
                kcu.column_name AS fk_column,
                ccu.table_name AS referenced_table,
                ccu.column_name AS referenced_column
            FROM information_schema.table_constraints AS tc
            JOIN information_schema.key_column_usage AS kcu
                ON tc.constraint_name = kcu.constraint_name
                AND tc.table_schema = kcu.table_schema
            JOIN information_schema.constraint_column_usage AS ccu
                ON ccu.constraint_name = tc.constraint_name
                AND ccu.table_schema = tc.table_schema
            WHERE tc.constraint_type = 'FOREIGN KEY'
                AND tc.table_name = :tabla
                AND tc.table_schema = 'public'
        """)
        
        result = session.execute(query_fk, {"tabla": tabla_nombre})
        fk_constraints = result.fetchall()
        
        # Validar cada foreign key
        for fk_col, ref_table, ref_col in fk_constraints:
            fk_col_lower = fk_col.lower()
            ref_table_lower = ref_table.lower()
            
            # Buscar el valor en los datos (case-insensitive)
            valor_fk = None
            for key, value in datos.items():
                if key.lower() == fk_col_lower:
                    valor_fk = value
                    break
            
            # Si hay un valor de FK y no es None, validar que existe
            if valor_fk is not None:
                try:
                    # Verificar que el valor existe en la tabla referenciada
                    check_query = text(f"""
                        SELECT COUNT(*) 
                        FROM public.{ref_table} 
                        WHERE {ref_col} = :valor
                    """)
                    count_result = session.execute(check_query, {"valor": valor_fk})
                    count = count_result.scalar()
                    
                    if count == 0:
                        errores.append({
                            'columna': fk_col,
                            'valor': valor_fk,
                            'tabla_referenciada': ref_table,
                            'columna_referenciada': ref_col,
                            'mensaje': f"Foreign key '{fk_col}' con valor '{valor_fk}' no existe en tabla '{ref_table}.{ref_col}'"
                        })
                except Exception as e:
                    logger.warning(f"Error validando FK {fk_col} -> {ref_table}.{ref_col}: {e}")
                    # No agregar error si la validación misma falla (puede ser problema de permisos)
        
    except Exception as e:
        logger.warning(f"Error obteniendo constraints de foreign key para tabla {tabla_nombre}: {e}")
        # Si no se pueden obtener los constraints, no validar (no fallar completamente)
    
    return errores


def validar_not_null(session, tabla_nombre: str, datos: Dict[str, Any], table) -> List[Dict[str, Any]]:
    """
    Valida que los datos tienen valores para todas las columnas NOT NULL.
    
    Args:
        session: Sesión de SQLAlchemy
        tabla_nombre: Nombre de la tabla
        datos: Diccionario con los datos a validar
        table: Objeto Table de SQLAlchemy
    
    Returns:
        List[Dict[str, Any]]: Lista de errores encontrados. Cada error tiene:
            - 'columna': nombre de la columna NOT NULL
            - 'mensaje': mensaje de error descriptivo
    """
    errores = []
    
    try:
        # Obtener columnas NOT NULL de la tabla
        for column in table.columns:
            # Verificar si la columna es NOT NULL y no tiene default
            is_not_null = not column.nullable
            has_default = column.server_default is not None or column.default is not None
            
            if is_not_null and not has_default:
                # Buscar el valor en los datos (case-insensitive)
                valor = None
                for key, value in datos.items():
                    if key.lower() == column.name.lower():
                        valor = value
                        break
                
                # Si el valor es None o string vacío, es un error
                if valor is None or (isinstance(valor, str) and valor.strip() == ''):
                    errores.append({
                        'columna': column.name,
                        'mensaje': f"Columna NOT NULL '{column.name}' no tiene valor o está vacía"
                    })
        
    except Exception as e:
        logger.warning(f"Error validando NOT NULL para tabla {tabla_nombre}: {e}")
    
    return errores


def convertir_valor_segun_tipo(valor: Any, tipo_columna) -> Any:
    """
    Convierte un valor según el tipo de columna de la base de datos.
    
    Args:
        valor: Valor a convertir
        tipo_columna: Tipo de columna de SQLAlchemy
    
    Returns:
        Valor convertido al tipo correcto
    """
    if valor is None:
        return None
    
    # Obtener el tipo Python del tipo SQLAlchemy
    tipo_python = tipo_columna.python_type if hasattr(tipo_columna, 'python_type') else type(valor)
    
    # Verificar también el nombre del tipo SQLAlchemy para detectar booleanos
    tipo_nombre = None
    if hasattr(tipo_columna, '__class__'):
        tipo_nombre = tipo_columna.__class__.__name__.lower()
    elif hasattr(tipo_columna, 'name'):
        tipo_nombre = str(tipo_columna.name).lower()
    
    # Verificar si es un tipo Boolean de SQLAlchemy
    es_booleano_sqlalchemy = False
    try:
        from sqlalchemy import Boolean
        if isinstance(tipo_columna, Boolean):
            es_booleano_sqlalchemy = True
    except:
        pass
    
    # Si ya es del tipo correcto, retornarlo
    if isinstance(valor, tipo_python):
        return valor
    
    # Convertir según el tipo
    try:
        # Booleanos: convertir strings 'true', 'false', '1', '0', etc.
        # Verificar tanto el tipo Python como el nombre del tipo SQLAlchemy
        es_booleano = (tipo_python == bool or 
                      es_booleano_sqlalchemy or
                      (tipo_nombre and 'bool' in tipo_nombre) or
                      (hasattr(tipo_columna, 'python_type') and tipo_columna.python_type == bool))
        
        if es_booleano:
            if isinstance(valor, bool):
                return valor
            if isinstance(valor, str):
                valor_str = valor.lower().strip()
                if valor_str in ('true', '1', 'yes', 'si', 'sí', 't', 'y', 'verdadero', 'v'):
                    return True
                elif valor_str in ('false', '0', 'no', 'n', 'f', 'falso'):
                    return False
                else:
                    # Si no se puede convertir, retornar False por defecto
                    logger.warning(f"No se pudo convertir '{valor}' a booleano, usando False")
                    return False
            if isinstance(valor, (int, float)):
                return bool(valor)
            # Si es None o vacío, retornar False por defecto para booleanos
            return False
        
        # Enteros
        elif tipo_python == int:
            if isinstance(valor, str):
                # Intentar convertir string a int
                try:
                    # Remover espacios y caracteres no numéricos
                    valor_limpio = valor.strip()
                    if valor_limpio:
                        return int(float(valor_limpio))  # Usar float primero para manejar "1.0"
                except (ValueError, TypeError):
                    logger.warning(f"No se pudo convertir '{valor}' a entero")
                    return None
            return int(valor)
        
        # Flotantes
        elif tipo_python == float:
            if isinstance(valor, str):
                try:
                    valor_limpio = valor.strip()
                    if valor_limpio:
                        return float(valor_limpio)
                except (ValueError, TypeError):
                    logger.warning(f"No se pudo convertir '{valor}' a flotante")
                    return None
            return float(valor)
        
        # Strings: asegurar que sea string
        elif tipo_python == str:
            return str(valor)
        
        # Fechas y timestamps
        elif hasattr(tipo_python, '__name__') and 'date' in tipo_python.__name__.lower():
            if isinstance(valor, str):
                # Intentar parsear fecha
                try:
                    from datetime import datetime
                    # Intentar varios formatos comunes (incluyendo microsegundos)
                    formatos = [
                        '%Y-%m-%d', 
                        '%d/%m/%Y', 
                        '%m/%d/%Y', 
                        '%Y-%m-%d %H:%M:%S',
                        '%Y-%m-%d %H:%M:%S.%f'  # Formato con microsegundos
                    ]
                    for fmt in formatos:
                        try:
                            return datetime.strptime(valor.strip(), fmt)
                        except ValueError:
                            continue
                    logger.warning(f"No se pudo parsear fecha '{valor}'")
                    return None
                except Exception as e:
                    logger.warning(f"Error parseando fecha '{valor}': {e}")
                    return None
            return valor
        
        # Por defecto, retornar el valor tal cual
        return valor
        
    except Exception as e:
        logger.warning(f"Error convirtiendo valor '{valor}' a tipo {tipo_python}: {e}")
        return valor

def import_one_file(session, model, ruta_archivo: str, formato: str, upsert: bool = False, keep_ids: bool = False) -> Tuple[int, int, Dict[str, Any]]:
    """
    Importa un archivo Excel/CSV a una tabla.
    
    Args:
        session: Sesión de SQLAlchemy
        model: Modelo de la tabla destino
        ruta_archivo: Ruta al archivo a importar
        formato: Formato del archivo ('xlsx' o 'csv')
        upsert: Si es True, actualiza registros existentes
        keep_ids: Si es True, mantiene los IDs del archivo
    
    Returns:
        Tuple[int, int, Dict[str, Any]]: (insertados, actualizados, errores_detalle)
        errores_detalle contiene:
            - 'errores_foreign_key': número de errores de FK
            - 'errores_not_null': número de errores NOT NULL
            - 'errores_tipo_dato': número de errores de tipo
            - 'errores_otros': número de otros errores
            - 'total_errores': total de errores
    """
    inserted = 0
    updated = 0
    errores_detalle = {
        'errores_foreign_key': 0,
        'errores_not_null': 0,
        'errores_tipo_dato': 0,
        'errores_otros': 0,
        'total_errores': 0
    }
    
    try:
        # Leer archivo
        if formato == 'xlsx':
            headers, rows = read_rows_from_xlsx(ruta_archivo)
        else:
            headers, rows = read_rows_from_csv(ruta_archivo)
        
        if not headers or not rows:
            logger.warning("Archivo vacío o sin datos")
            return 0, 0
        
        # Normalizar headers
        headers = normalize_header_names(headers)
        
        # Obtener columnas del modelo con sus tipos
        columnas_modelo = {c.name.lower(): c for c in model.__table__.columns}
        
        # Mapear headers a columnas del modelo (incluyendo el tipo)
        mapeo_columnas = {}
        tipos_columnas = {}  # Almacenar tipos de columnas
        for header in headers:
            if header.lower() in columnas_modelo:
                columna_modelo = columnas_modelo[header.lower()]
                nombre_columna = columna_modelo.name
                mapeo_columnas[header] = nombre_columna
                tipos_columnas[nombre_columna] = columna_modelo.type  # Guardar el tipo
                # Log de depuración para columnas booleanas
                tipo_col = columna_modelo.type
                es_booleano = False
                if hasattr(tipo_col, 'python_type') and tipo_col.python_type == bool:
                    es_booleano = True
                    logger.debug(f"Columna booleana detectada: '{nombre_columna}' (tipo: {tipo_col})")
                elif hasattr(tipo_col, '__class__'):
                    tipo_nombre = tipo_col.__class__.__name__.lower()
                    if 'bool' in tipo_nombre:
                        es_booleano = True
                        logger.debug(f"Columna booleana detectada por nombre: '{nombre_columna}' (tipo: {tipo_nombre})")
                # Verificar también el tipo de datos de PostgreSQL
                if hasattr(tipo_col, 'type_affinity'):
                    try:
                        from sqlalchemy import Boolean
                        if isinstance(tipo_col, Boolean):
                            es_booleano = True
                            logger.debug(f"Columna booleana detectada por tipo SQLAlchemy: '{nombre_columna}'")
                    except:
                        pass
        
        if not mapeo_columnas:
            logger.error("No se encontraron columnas coincidentes")
            return 0, 0
        
        # Procesar filas y preparar datos para bulk insert
        filas_datos = []
        for fila in rows:
            try:
                datos = {}
                for i, valor in enumerate(fila):
                    if i < len(headers) and headers[i] in mapeo_columnas:
                        columna = mapeo_columnas[headers[i]]
                        # Convertir valor según el tipo de columna
                        tipo_columna = tipos_columnas.get(columna)
                        if tipo_columna:
                            try:
                                valor_convertido = convertir_valor_segun_tipo(valor, tipo_columna)
                                # Log de depuración para booleanos
                                if hasattr(tipo_columna, 'python_type') and tipo_columna.python_type == bool:
                                    logger.debug(f"Columna '{columna}': '{valor}' ({type(valor).__name__}) -> {valor_convertido} ({type(valor_convertido).__name__})")
                            except Exception as e:
                                logger.warning(f"Error convirtiendo valor '{valor}' para columna '{columna}': {e}")
                                valor_convertido = valor
                        else:
                            valor_convertido = valor
                        datos[columna] = valor_convertido
                
                if datos:
                    filas_datos.append((len(filas_datos) + 1, datos))  # Guardar índice de fila junto con datos
            except Exception as e:
                logger.warning(f"Error procesando fila: {e}")
                errores_detalle['errores_otros'] += 1
                continue
        
        # Usar SQLAlchemy Core directamente pero con el mapper
        # Obtener la tabla de forma segura usando caché para evitar configurar relaciones
        from sqlalchemy import insert, select, update, MetaData, text
        
        if filas_datos:
            try:
                # Obtener nombre de tabla sin acceder al mapper
                tabla_nombre = obtener_nombre_tabla_seguro(model)
                
                # Usar caché para obtener la tabla sin disparar configuración del mapper
                if tabla_nombre not in _TABLES_CACHE:
                    # Obtener tabla directamente desde la metadata reflejada
                    # Esto evita acceder al mapper que puede disparar configuración de relaciones
                    metadata = MetaData()
                    metadata.reflect(bind=session.bind, only=[tabla_nombre])
                    _TABLES_CACHE[tabla_nombre] = metadata.tables[tabla_nombre]
                
                table = _TABLES_CACHE[tabla_nombre]
                
                # Contadores de errores por tipo
                errores_foreign_key = 0
                errores_not_null = 0
                errores_tipo_dato = 0
                errores_otros = 0
                
                if upsert:
                    # Para upsert, necesitamos procesar individualmente
                    # Usar savepoints para aislar errores por fila sin abortar toda la transacción
                    commit_interval = 100  # Commit cada 100 filas
                    
                    for fila_idx, datos in filas_datos:
                        savepoint_name = f"sp_fila_{fila_idx}"
                        try:
                            # Validar foreign keys y NOT NULL antes de procesar
                            errores_fk = validar_foreign_keys(session, tabla_nombre, datos)
                            errores_nn = validar_not_null(session, tabla_nombre, datos, table)
                            
                            # Si hay errores de validación, reportarlos y saltar esta fila
                            if errores_fk or errores_nn:
                                for error in errores_fk:
                                    logger.warning(f"Fila {fila_idx} - Validación FK fallida: {error['mensaje']}")
                                    errores_foreign_key += 1
                                for error in errores_nn:
                                    logger.warning(f"Fila {fila_idx} - Validación NOT NULL fallida: {error['mensaje']}")
                                    errores_not_null += 1
                                # Log datos de la fila para diagnóstico
                                datos_sanitizados = {k: ('***' if 'password' in k.lower() or 'pass' in k.lower() else v) 
                                                    for k, v in datos.items()}
                                logger.debug(f"Fila {fila_idx} - Datos con error de validación: {datos_sanitizados}")
                                # No procesar esta fila si tiene errores de validación
                                continue
                            
                            # Crear savepoint antes de procesar cada fila
                            # Esto permite hacer rollback solo de esta fila si hay error
                            session.execute(text(f"SAVEPOINT {savepoint_name}"))
                            
                            # Buscar clave primaria
                            pk_column = None
                            for col in table.primary_key.columns:
                                if col.name in datos:
                                    pk_column = col.name
                                    break
                            
                            if pk_column and datos.get(pk_column):
                                # Verificar si existe usando select directo (Core, no ORM)
                                stmt_select = select(table).where(
                                    table.columns[pk_column] == datos[pk_column]
                                )
                                result = session.execute(stmt_select).first()
                                
                                if result:
                                    # Actualizar usando update directo (Core, no ORM)
                                    stmt_update = update(table).where(
                                        table.columns[pk_column] == datos[pk_column]
                                    ).values(**datos)
                                    session.execute(stmt_update)
                                    updated += 1
                                else:
                                    # Insertar usando insert directo (Core, no ORM)
                                    stmt_insert = insert(table).values(**datos)
                                    session.execute(stmt_insert)
                                    inserted += 1
                            else:
                                # No hay PK, insertar directamente
                                stmt_insert = insert(table).values(**datos)
                                session.execute(stmt_insert)
                                inserted += 1
                            
                            # Liberar savepoint si todo salió bien
                            session.execute(text(f"RELEASE SAVEPOINT {savepoint_name}"))
                            
                            # Commit periódico para evitar perder trabajo y mantener transacción válida
                            if fila_idx % commit_interval == 0:
                                session.commit()
                                logger.debug(f"Commit periódico: {fila_idx} filas procesadas")
                                
                        except Exception as e:
                            # Determinar tipo de error
                            error_msg = str(e).lower()
                            tipo_error = "otro"
                            if 'foreign key' in error_msg or 'violates foreign key' in error_msg:
                                errores_foreign_key += 1
                                tipo_error = "foreign_key"
                            elif 'not null' in error_msg or 'null value' in error_msg:
                                errores_not_null += 1
                                tipo_error = "not_null"
                            elif 'invalid input' in error_msg or 'type' in error_msg or 'invalid' in error_msg:
                                errores_tipo_dato += 1
                                tipo_error = "tipo_dato"
                            else:
                                errores_otros += 1
                            
                            # Log detallado del error
                            logger.warning(f"Fila {fila_idx} - Error tipo '{tipo_error}': {e}")
                            # Log datos de la fila (sanitizado, sin valores sensibles)
                            datos_sanitizados = {k: ('***' if 'password' in k.lower() or 'pass' in k.lower() else v) 
                                                for k, v in datos.items()}
                            logger.debug(f"Fila {fila_idx} - Datos: {datos_sanitizados}")
                            
                            # Hacer rollback al savepoint (solo esta fila)
                            try:
                                session.execute(text(f"ROLLBACK TO SAVEPOINT {savepoint_name}"))
                                logger.debug(f"Fila {fila_idx} - Rollback al savepoint realizado exitosamente")
                            except Exception as rollback_error:
                                # Si el rollback al savepoint falla, hacer rollback completo
                                logger.error(f"Fila {fila_idx} - Error haciendo rollback al savepoint {savepoint_name}: {rollback_error}. Haciendo rollback completo.")
                                try:
                                    session.rollback()
                                except Exception as full_rollback_error:
                                    logger.error(f"Fila {fila_idx} - Error crítico en rollback completo: {full_rollback_error}")
                            continue
                else:
                    # Usar insert() statements directamente (Core, no ORM)
                    # Insertar en lotes para mejor rendimiento
                    # Primero validar y filtrar filas inválidas
                    batch_size = 1000
                    filas_validas = []
                    
                    for fila_idx, datos in filas_datos:
                        # Validar foreign keys y NOT NULL antes de agregar al batch
                        errores_fk = validar_foreign_keys(session, tabla_nombre, datos)
                        errores_nn = validar_not_null(session, tabla_nombre, datos, table)
                        
                        if errores_fk or errores_nn:
                            # Reportar errores y no incluir esta fila
                            for error in errores_fk:
                                logger.warning(f"Fila {fila_idx} - Validación FK fallida: {error['mensaje']}")
                                errores_foreign_key += 1
                            for error in errores_nn:
                                logger.warning(f"Fila {fila_idx} - Validación NOT NULL fallida: {error['mensaje']}")
                                errores_not_null += 1
                            # Log datos de la fila para diagnóstico
                            datos_sanitizados = {k: ('***' if 'password' in k.lower() or 'pass' in k.lower() else v) 
                                                for k, v in datos.items()}
                            logger.debug(f"Fila {fila_idx} - Datos con error de validación: {datos_sanitizados}")
                            continue
                        
                        # Agregar solo los datos (sin el índice) a la lista de filas válidas
                        filas_validas.append(datos)
                    
                    # Insertar filas válidas en lotes
                    for i in range(0, len(filas_validas), batch_size):
                        batch = filas_validas[i:i + batch_size]
                        # Filtrar None de las claves primarias para evitar SAWarning
                        # Si la PK tiene autoincrement en la BD, no necesita valor explícito
                        batch_clean = []
                        for row in batch:
                            row_clean = row.copy()
                            # Remover valores None de columnas PK que puedan tener autoincrement
                            for pk_col in table.primary_key.columns:
                                if pk_col.name in row_clean and row_clean[pk_col.name] is None:
                                    # Si la columna PK es None y puede tener autoincrement, removerla
                                    # La BD generará el valor automáticamente
                                    if pk_col.autoincrement or pk_col.server_default is not None:
                                        del row_clean[pk_col.name]
                            batch_clean.append(row_clean)
                        
                        try:
                            stmt = insert(table).values(batch_clean)
                            result = session.execute(stmt)
                            inserted += len(batch_clean)
                        except Exception as batch_error:
                            # Si un batch falla, intentar insertar fila por fila para identificar el problema
                            error_msg = str(batch_error).lower()
                            if 'foreign key' in error_msg or 'violates foreign key' in error_msg:
                                errores_foreign_key += len(batch_clean)
                            elif 'not null' in error_msg or 'null value' in error_msg:
                                errores_not_null += len(batch_clean)
                            elif 'invalid input' in error_msg or 'type' in error_msg:
                                errores_tipo_dato += len(batch_clean)
                            else:
                                errores_otros += len(batch_clean)
                            logger.warning(f"Error insertando batch: {batch_error}")
                            # Continuar con el siguiente batch
                            continue
                
                session.commit()
                
                # Actualizar contadores de errores en el diccionario de retorno
                errores_detalle['errores_foreign_key'] = errores_foreign_key
                errores_detalle['errores_not_null'] = errores_not_null
                errores_detalle['errores_tipo_dato'] = errores_tipo_dato
                errores_detalle['errores_otros'] = errores_otros
                errores_detalle['total_errores'] = errores_foreign_key + errores_not_null + errores_tipo_dato + errores_otros
                
                # Log resumen de errores
                if errores_detalle['total_errores'] > 0:
                    logger.warning(f"Importación completada con errores: {inserted} insertados, {updated} actualizados")
                    logger.warning(f"Errores encontrados: FK={errores_foreign_key}, NOT NULL={errores_not_null}, Tipo={errores_tipo_dato}, Otros={errores_otros}")
                else:
                    logger.info(f"Importación completada: {inserted} insertados, {updated} actualizados")
            except Exception as e:
                logger.error(f"Error en insert: {e}")
                session.rollback()
                raise
        else:
            logger.warning("No hay datos para importar")
        
    except Exception as e:
        logger.error(f"Error importando archivo: {e}")
        session.rollback()
        raise
    
    return inserted, updated, errores_detalle


def _detectar_modelo_para_archivo(ruta_archivo: str, formato: str, tabla_cli: Optional[str]) -> Tuple[str, Any]:
    """
    Determina el modelo de destino para un archivo dado.
    Usa una estrategia en cascada:
    1. Tabla especificada explícitamente (tabla_cli)
    2. Detección por nombre de archivo (mapeo)
    3. Detección por tipo de análisis (contenido)
    4. Detección por columnas del archivo

    Args:
        ruta_archivo: Ruta del archivo a importar.
        formato: Formato detectado del archivo (csv/xlsx).
        tabla_cli: Tabla proporcionada explícitamente por CLI.

    Returns:
        Tuple con (nombre_tabla, modelo_sqlalchemy).

    Raises:
        ValueError: Si no se puede determinar la tabla destino.
    """
    # Estrategia 1: Tabla especificada explícitamente
    if tabla_cli:
        try:
            modelo = obtener_modelo(tabla_cli)
            logger.info(f"Usando tabla especificada explícitamente: '{tabla_cli}'")
            return tabla_cli, modelo
        except Exception as exc:
            raise ValueError(f"No se pudo obtener el modelo para la tabla '{tabla_cli}': {exc}") from exc

    # Estrategia 2: Detección por nombre de archivo
    nombre_archivo = os.path.basename(ruta_archivo)
    tabla_por_nombre = detectar_tabla_por_nombre_archivo(nombre_archivo)
    if tabla_por_nombre:
        try:
            modelo = obtener_modelo(tabla_por_nombre)
            logger.info(f"Tabla detectada por nombre de archivo: '{tabla_por_nombre}' (archivo: '{nombre_archivo}')")
            return tabla_por_nombre, modelo
        except Exception as exc:
            logger.warning(f"Tabla detectada por nombre '{tabla_por_nombre}' pero no se pudo obtener modelo: {exc}. Continuando con otras estrategias...")

    # Estrategia 3: Detección por tipo de análisis (contenido)
    try:
        tipo_analisis = detectar_tipo_analisis_por_contenido(ruta_archivo)
        if tipo_analisis:
            logger.info(f"Tipo de análisis detectado por contenido: '{tipo_analisis}'")
            # Si el tipo de análisis coincide con una tabla, usarla
            if tipo_analisis in MODELS:
                try:
                    modelo = obtener_modelo(tipo_analisis)
                    logger.info(f"Tabla detectada por tipo de análisis: '{tipo_analisis}'")
                    return tipo_analisis, modelo
                except Exception as exc:
                    logger.warning(f"Tipo de análisis '{tipo_analisis}' detectado pero no se pudo obtener modelo: {exc}")
    except Exception as exc:
        logger.debug(f"Error detectando tipo de análisis por contenido: {exc}")
        tipo_analisis = None

    # Estrategia 4: Detección por columnas del archivo
    try:
        if formato == 'xlsx':
            headers, _ = read_rows_from_xlsx(ruta_archivo, max_rows=1)
        else:
            headers, _ = read_rows_from_csv(ruta_archivo, max_rows=1)
    except Exception as exc:
        raise ValueError(f"No se pudieron leer encabezados del archivo: {exc}") from exc

    if not headers:
        # Generar mensaje de error más descriptivo
        sugerencias = []
        if nombre_archivo:
            nombre_sin_ext = os.path.splitext(nombre_archivo)[0].lower()
            # Buscar tablas similares
            tablas_similares = [t for t in MODELS.keys() if nombre_sin_ext in t or t in nombre_sin_ext]
            if tablas_similares:
                sugerencias = tablas_similares[:5]
        
        mensaje_error = "El archivo no contiene encabezados para detectar la tabla destino."
        if sugerencias:
            mensaje_error += f" Tablas sugeridas basadas en el nombre del archivo: {', '.join(sugerencias)}"
        raise ValueError(mensaje_error)

    tabla_detectada = detectar_tabla_por_columnas(None, headers, tipo_analisis)
    if not tabla_detectada:
        # Generar mensaje de error más descriptivo con sugerencias
        sugerencias = []
        
        # Buscar tablas con columnas similares
        headers_lower = [h.lower() for h in headers if h]
        mejor_coincidencia = None
        mejor_puntuacion = 0
        
        for tabla_nombre, model in MODELS.items():
            try:
                columnas_modelo = [c.name.lower() for c in model.__table__.columns]
                coincidencias = sum(1 for h in headers_lower if h in columnas_modelo)
                puntuacion = coincidencias / len(headers_lower) if headers_lower else 0
                
                if puntuacion > mejor_puntuacion:
                    mejor_puntuacion = puntuacion
                    mejor_coincidencia = tabla_nombre
                
                # Agregar a sugerencias si tiene al menos 20% de coincidencia
                if puntuacion >= 0.2:
                    sugerencias.append((tabla_nombre, puntuacion))
            except Exception:
                continue
        
        # Ordenar sugerencias por puntuación
        sugerencias.sort(key=lambda x: x[1], reverse=True)
        sugerencias_nombres = [t[0] for t in sugerencias[:5]]
        
        mensaje_error = f"No se pudo detectar automáticamente la tabla destino. Columnas encontradas: {', '.join(headers[:10])}"
        if sugerencias_nombres:
            mensaje_error += f" Tablas sugeridas (por similitud de columnas): {', '.join(sugerencias_nombres)}"
        if nombre_archivo:
            nombre_sin_ext = os.path.splitext(nombre_archivo)[0].lower()
            tablas_por_nombre = [t for t in MODELS.keys() if nombre_sin_ext in t or t in nombre_sin_ext]
            if tablas_por_nombre:
                mensaje_error += f" Tablas sugeridas (por nombre de archivo): {', '.join(tablas_por_nombre[:3])}"
        mensaje_error += " Especifica una tabla con '--table' o renombra el archivo para que coincida con el nombre de la tabla."
        
        raise ValueError(mensaje_error)

    try:
        modelo = obtener_modelo(tabla_detectada)
        logger.info(f"Tabla detectada por columnas: '{tabla_detectada}'")
    except Exception as exc:
        raise ValueError(f"No se pudo obtener el modelo para la tabla detectada '{tabla_detectada}': {exc}") from exc

    return tabla_detectada, modelo


def _procesar_archivo_cli(
    SessionFactory,
    ruta_archivo: str,
    tabla_cli: Optional[str],
    upsert: bool,
    keep_ids: bool,
) -> Dict[str, Any]:
    """Procesa un archivo desde CLI y retorna el resultado."""
    resultado = {
        "archivo": ruta_archivo,
        "tabla": None,
        "formato": None,
        "insertados": 0,
        "actualizados": 0,
        "upsert": upsert,
        "keep_ids": keep_ids,
        "exito": False,
        "mensaje": "",
    }

    if not os.path.exists(ruta_archivo):
        resultado["mensaje"] = "El archivo no existe"
        return resultado

    formato = detect_format_from_path(ruta_archivo)
    if not formato:
        resultado["mensaje"] = "Formato de archivo no soportado (use .csv, .xlsx o .xls)"
        return resultado

    resultado["formato"] = formato

    try:
        tabla_destino, modelo = _detectar_modelo_para_archivo(ruta_archivo, formato, tabla_cli)
        resultado["tabla"] = tabla_destino
    except ValueError as exc:
        resultado["mensaje"] = str(exc)
        return resultado

    session = SessionFactory()
    try:
        inserted, updated, errores_detalle = import_one_file(session, modelo, ruta_archivo, formato, upsert=upsert, keep_ids=keep_ids)
        resultado["insertados"] = inserted
        resultado["actualizados"] = updated
        resultado["errores"] = errores_detalle
        resultado["exito"] = True
        resultado["mensaje"] = "Importación completada"
    except Exception as exc:
        logger.error(f"Error importando '{ruta_archivo}': {exc}", exc_info=True)
        resultado["mensaje"] = f"Error importando archivo: {exc}"
    finally:
        session.close()

    return resultado


def main():
    """Entrada CLI para importar uno o varios archivos."""
    parser = argparse.ArgumentParser(
        description="Importa uno o varios archivos CSV/XLSX a la base de datos del proyecto INIA."
    )
    parser.add_argument(
        "files",
        nargs="+",
        help="Ruta(s) de archivos a importar (.csv, .xlsx, .xls).",
    )
    parser.add_argument(
        "--table",
        help="Nombre de la tabla destino. Si no se especifica se intentará detectar automáticamente.",
    )
    parser.add_argument(
        "--upsert",
        action="store_true",
        help="Actualiza los registros existentes coincidencia por clave primaria.",
    )
    parser.add_argument(
        "--keep-ids",
        action="store_true",
        help="Mantiene los IDs provistos en el archivo en lugar de generar nuevos.",
    )
    args = parser.parse_args()

    try:
        engine = obtener_engine()
        inicializar_automap(engine)
    except Exception as exc:
        logger.error(f"No se pudo inicializar la conexión a la base de datos: {exc}", exc_info=True)
        raise SystemExit(1) from exc

    SessionFactory = sessionmaker(bind=engine)

    resultados: List[Dict[str, Any]] = []
    for ruta in args.files:
        logger.info(f"Procesando archivo '{ruta}'...")
        resultado = _procesar_archivo_cli(SessionFactory, ruta, args.table, args.upsert, args.keep_ids)
        resultados.append(resultado)
        if resultado["exito"]:
            logger.info(
                f"  -> Tabla: {resultado['tabla']} | Insertados: {resultado['insertados']} | Actualizados: {resultado['actualizados']}"
            )
        else:
            logger.warning(f"  -> Falló: {resultado['mensaje']}")

    exitosos = [r for r in resultados if r["exito"]]
    fallidos = [r for r in resultados if not r["exito"]]

    print("\nResumen importación:")
    for res in resultados:
        estado = "OK" if res["exito"] else "ERROR"
        detalle = (
            f"{res['insertados']} insertados / {res['actualizados']} actualizados"
            if res["exito"]
            else res["mensaje"]
        )
        tabla = res["tabla"] or "desconocida"
        print(f" - [{estado}] {res['archivo']} -> tabla '{tabla}' ({detalle})")

    print(
        f"\nArchivos procesados: {len(resultados)} | Exitosos: {len(exitosos)} | Con error: {len(fallidos)}"
    )

    if fallidos:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
