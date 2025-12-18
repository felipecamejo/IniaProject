import os
import csv
import argparse
import logging
from datetime import date, datetime
from typing import Optional, Dict, Any, List
from urllib.parse import quote_plus

# Importar dependencias usando módulo común
from dependencies_common import importar_sqlalchemy, importar_openpyxl

# Importar SQLAlchemy
create_engine, text, inspect, _, sessionmaker, automap_base = importar_sqlalchemy()

# Importar openpyxl
OPENPYXL_AVAILABLE, Workbook, load_workbook, get_column_letter, Font, PatternFill, Border, Side, Alignment = importar_openpyxl()
import openpyxl.styles

# Configuración de logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Importar funciones comunes de base de datos
from db_common import (
    obtener_engine,
    inicializar_automap,
    obtener_modelo,
    obtener_nombre_tabla_seguro,
    Base,
    MODELS
)

# ================================
# MÓDULO: LOGGING Y UTILIDADES
# ================================
def log_ok(message: str):
    """Log de éxito."""
    logger.info(f"✓ {message}")

def log_fail(message: str):
    """Log de error."""
    logger.error(f"✗ {message}")

def log_step(message: str):
    """Log de paso."""
    logger.info(f"→ {message}")

def ensure_output_dir(path: str) -> str:
    """Asegura que el directorio de salida exista."""
    os.makedirs(path, exist_ok=True)
    return os.path.abspath(path)

def serialize_value(value):
    """Serializa un valor para exportación."""
    # #region agent log
    import json as _json_log
    try:
        _log_data = {"value_type": str(type(value).__name__), "value_repr": repr(value)[:200] if value is not None else "None"}
        with open(r"c:\Github\IniaProject\.cursor\debug.log", "a", encoding="utf-8") as _f:
            _f.write(_json_log.dumps({"location": "ExportExcel.py:serialize_value", "message": "serialize_value called", "data": _log_data, "timestamp": __import__('time').time(), "sessionId": "debug-session", "hypothesisId": "B"}) + "\n")
    except: pass
    # #endregion
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=" ")
    if isinstance(value, bool):
        return "true" if value else "false"
    # #region agent log
    if isinstance(value, str):
        try:
            _encoded = value.encode('utf-8')
            with open(r"c:\Github\IniaProject\.cursor\debug.log", "a", encoding="utf-8") as _f:
                _f.write(_json_log.dumps({"location": "ExportExcel.py:serialize_value_str", "message": "string value OK", "data": {"value_preview": value[:100] if len(value) > 100 else value}, "timestamp": __import__('time').time(), "sessionId": "debug-session", "hypothesisId": "B"}) + "\n")
        except UnicodeEncodeError as _ue:
            with open(r"c:\Github\IniaProject\.cursor\debug.log", "a", encoding="utf-8") as _f:
                _f.write(_json_log.dumps({"location": "ExportExcel.py:serialize_value_str_error", "message": "UNICODE ERROR in string", "data": {"error": str(_ue), "value_repr": repr(value)[:200]}, "timestamp": __import__('time').time(), "sessionId": "debug-session", "hypothesisId": "B"}) + "\n")
        except: pass
    # #endregion
    return value

# ================================
# MÓDULO: VALIDACIÓN DE COLUMNAS
# ================================
def verificar_estructura_tabla(session, tabla_nombre: str) -> list:
    """Verifica la estructura real de una tabla en la base de datos."""
    try:
        query = text("""
            SELECT column_name, data_type, is_nullable
            FROM information_schema.columns
            WHERE table_schema = 'public' AND table_name = :tabla
            ORDER BY ordinal_position
        """)
        result = session.execute(query, {"tabla": tabla_nombre}).fetchall()
        return [row[0] for row in result]
    except Exception as e:
        log_fail(f"Error verificando estructura de {tabla_nombre}: {e}")
        return []

def tiene_primary_key(session, tabla_nombre: str) -> bool:
    """Verifica si una tabla tiene Primary Key."""
    try:
        query = text("""
            SELECT COUNT(*)
            FROM information_schema.table_constraints tc
            JOIN information_schema.constraint_column_usage ccu
                ON tc.constraint_name = ccu.constraint_name
                AND tc.table_schema = ccu.table_schema
            WHERE tc.constraint_type = 'PRIMARY KEY'
                AND tc.table_schema = 'public'
                AND tc.table_name = :tabla
        """)
        result = session.execute(query, {"tabla": tabla_nombre}).fetchone()
        return result[0] > 0 if result else False
    except Exception as e:
        log_fail(f"Error verificando PK de {tabla_nombre}: {e}")
        return False

def obtener_tablas_sin_pk(session) -> list:
    """Obtiene todas las tablas que no tienen Primary Key."""
    try:
        query = text("""
            SELECT t.table_name
            FROM information_schema.tables t
            WHERE t.table_schema = 'public'
                AND t.table_type = 'BASE TABLE'
                AND NOT EXISTS (
                    SELECT 1
                    FROM information_schema.table_constraints tc
                    WHERE tc.table_schema = 'public'
                        AND tc.table_name = t.table_name
                        AND tc.constraint_type = 'PRIMARY KEY'
                )
            ORDER BY t.table_name
        """)
        result = session.execute(query).fetchall()
        tablas_sin_pk = [row[0] for row in result]
        if tablas_sin_pk:
            log_step(f"Tablas sin PK encontradas: {len(tablas_sin_pk)} - {', '.join(tablas_sin_pk)}")
        return tablas_sin_pk
    except Exception as e:
        log_fail(f"Error obteniendo tablas sin PK: {e}")
        return []

def obtener_tablas_relacionadas(session, tabla_principal: str) -> list:
    """Obtiene tablas relacionadas (sin PK) que están vinculadas a una tabla principal mediante Foreign Keys."""
    try:
        # Usar referential_constraints para obtener las tablas referenciadas correctamente
        # En PostgreSQL, usamos constraint_column_usage para obtener la tabla referenciada
        query = text("""
            SELECT DISTINCT tc.table_name
            FROM information_schema.table_constraints tc
            JOIN information_schema.referential_constraints rc
                ON tc.constraint_name = rc.constraint_name
                AND tc.table_schema = rc.constraint_schema
            JOIN information_schema.constraint_column_usage ccu
                ON rc.unique_constraint_name = ccu.constraint_name
                AND rc.unique_constraint_schema = ccu.constraint_schema
            WHERE tc.table_schema = 'public'
                AND tc.constraint_type = 'FOREIGN KEY'
                AND rc.unique_constraint_schema = 'public'
                AND ccu.table_name = :tabla_principal
                AND tc.table_name != :tabla_principal
                AND NOT EXISTS (
                    SELECT 1
                    FROM information_schema.table_constraints pk_tc
                    WHERE pk_tc.table_schema = 'public'
                        AND pk_tc.table_name = tc.table_name
                        AND pk_tc.constraint_type = 'PRIMARY KEY'
                )
            ORDER BY tc.table_name
        """)
        result = session.execute(query, {"tabla_principal": tabla_principal}).fetchall()
        tablas_relacionadas = [row[0] for row in result]
        if tablas_relacionadas:
            log_step(f"Tablas relacionadas con {tabla_principal} (sin PK): {len(tablas_relacionadas)} - {', '.join(tablas_relacionadas)}")
        return tablas_relacionadas
    except Exception as e:
        log_fail(f"Error obteniendo tablas relacionadas con {tabla_principal}: {e}")
        # Hacer rollback para que la sesión pueda continuar
        try:
            session.rollback()
        except:
            pass
        return []

def filtrar_columnas_analisis(columnas: list, tabla: str) -> list:
    """Filtra columnas para exportar solo datos de análisis, excluyendo campos administrativos."""
    campos_excluir = {
        'id', 'dosn_id', 'germinacion_id', 'pms_id', 'pureza_id', 
        'pureza_pnotatum_id', 'sanitario_id', 'tetrazolio_id', 'recibo_id',
        'activo', 'dosn_activo', 'germinacion_activo', 'pms_activo', 
        'pureza_activo', 'sanitario_activo', 'tetrazolio_activo', 'recibo_activo',
        'repetido', 'dosn_repetido', 'germinacion_repetido', 'pms_repetido',
        'pureza_repetido', 'sanitario_repetido', 'tetrazolio_repetido',
        'fecha_creacion', 'dosn_fecha_creacion', 'germinacion_fecha_creacion',
        'pms_fecha_creacion', 'pureza_fecha_creacion', 'sanitario_fechacreacion',
        'tetrazolio_fecha_creacion', 'fecha_repeticion', 'dosn_fecha_repeticion',
        'germinacion_fecha_repeticion', 'pms_fecha_repeticion', 'pureza_fecha_repeticion',
        'sanitario_fecharepeticion', 'tetrazolio_fecha_repeticion',
        'deposito_id'
    }
    
    columnas_analisis = [col for col in columnas if col not in campos_excluir]
    log_step(f"Filtrando {tabla}: {len(columnas)} → {len(columnas_analisis)} columnas de análisis")
    return columnas_analisis

def obtener_nombre_tabla(model) -> str:
    """Obtiene el nombre de la tabla de un modelo."""
    # Intentar diferentes métodos para obtener el nombre de la tabla
    try:
        if hasattr(model, '__tablename__'):
            return model.__tablename__
    except:
        pass
    
    try:
        if hasattr(model, '__table__') and hasattr(model.__table__, 'name'):
            return model.__table__.name
    except:
        pass
    
    # Si no se puede obtener, usar el nombre de la clase
    return model.__name__.lower()

def obtener_columnas_validas(session, model) -> tuple:
    """Obtiene y valida las columnas de una tabla. Retorna (columnas_validas, columnas_analisis)."""
    table = obtener_nombre_tabla(model)
    columns = [c.name for c in model.__table__.columns]
    
    columnas_reales = verificar_estructura_tabla(session, table)
    if not columnas_reales:
        log_fail(f"No se pudo verificar estructura de {table}")
        return [], []
    
    columnas_validas = [col for col in columns if col in columnas_reales]
    columnas_faltantes = set(columns) - set(columnas_validas)
    
    if columnas_faltantes:
        log_fail(f"Columnas faltantes en {table}: {columnas_faltantes}")
        log_step(f"Exportando solo columnas existentes: {columnas_validas}")
    
    if not columnas_validas:
        log_fail(f"No hay columnas válidas para exportar en {table}")
        return [], []
    
    columnas_analisis = filtrar_columnas_analisis(columnas_validas, table)
    return columnas_validas, columnas_analisis

def obtener_datos_tabla(session, tabla_nombre: str, columnas: list, filtro_where: Optional[str] = None, filtro_params: Optional[Dict[str, Any]] = None) -> list:
    """
    Obtiene los datos de una tabla usando SQL directo con filtros opcionales.
    
    Args:
        session: Sesión de SQLAlchemy
        tabla_nombre: Nombre de la tabla
        columnas: Lista de columnas a seleccionar
        filtro_where: Cláusula WHERE adicional (opcional)
        filtro_params: Parámetros para la cláusula WHERE (opcional)
    
    Returns:
        Lista de filas de resultados
    """
    # #region agent log
    import json as _json_log
    try:
        with open(r"c:\Github\IniaProject\.cursor\debug.log", "a", encoding="utf-8") as _f:
            _f.write(_json_log.dumps({"location": "ExportExcel.py:obtener_datos_tabla", "message": "Query start", "data": {"tabla": tabla_nombre, "columnas": columnas[:5], "filtro": filtro_where}, "timestamp": __import__('time').time(), "sessionId": "debug-session", "hypothesisId": "C"}) + "\n")
    except: pass
    # #endregion
    query_str = f"SELECT {', '.join(columnas)} FROM {tabla_nombre}"
    if filtro_where:
        query_str += f" WHERE {filtro_where}"
    
    query = text(query_str)
    if filtro_params:
        result = session.execute(query, filtro_params)
    else:
        result = session.execute(query)
    rows = result.fetchall()
    # #region agent log
    try:
        with open(r"c:\Github\IniaProject\.cursor\debug.log", "a", encoding="utf-8") as _f:
            _f.write(_json_log.dumps({"location": "ExportExcel.py:obtener_datos_tabla_result", "message": "Query result", "data": {"tabla": tabla_nombre, "row_count": len(rows)}, "timestamp": __import__('time').time(), "sessionId": "debug-session", "hypothesisId": "C"}) + "\n")
        # Log first row data types and sample values
        if rows:
            first_row_info = []
            for i, val in enumerate(rows[0]):
                col_info = {"col_idx": i, "type": str(type(val).__name__)}
                if isinstance(val, str):
                    try:
                        val.encode('utf-8')
                        col_info["utf8_ok"] = True
                        col_info["preview"] = val[:50] if len(val) > 50 else val
                    except UnicodeEncodeError as ue:
                        col_info["utf8_ok"] = False
                        col_info["error"] = str(ue)
                        col_info["repr"] = repr(val)[:100]
                elif val is not None:
                    col_info["repr"] = repr(val)[:50]
                first_row_info.append(col_info)
            with open(r"c:\Github\IniaProject\.cursor\debug.log", "a", encoding="utf-8") as _f:
                _f.write(_json_log.dumps({"location": "ExportExcel.py:obtener_datos_tabla_first_row", "message": "First row analysis", "data": {"tabla": tabla_nombre, "columns_info": first_row_info}, "timestamp": __import__('time').time(), "sessionId": "debug-session", "hypothesisId": "A"}) + "\n")
    except Exception as _e:
        try:
            with open(r"c:\Github\IniaProject\.cursor\debug.log", "a", encoding="utf-8") as _f:
                _f.write(_json_log.dumps({"location": "ExportExcel.py:obtener_datos_tabla_log_error", "message": "Error logging row data", "data": {"error": str(_e)}, "timestamp": __import__('time').time(), "sessionId": "debug-session", "hypothesisId": "A"}) + "\n")
        except: pass
    # #endregion
    return rows

# ================================
# MÓDULO: MAPEO DE CAMPOS DE FECHA
# ================================
def obtener_campo_fecha_analisis(tipo_analisis: str, campo_fecha: Optional[str] = None) -> Optional[str]:
    """
    Obtiene el campo de fecha de análisis apropiado para un tipo de análisis.
    
    Args:
        tipo_analisis: Tipo de análisis ('dosn', 'pureza', 'germinacion', 'pms', 'sanitario', 'tetrazolio', 'pureza_pnotatum')
        campo_fecha: Campo de fecha específico solicitado. Si es 'auto' o None, usa detección automática.
    
    Returns:
        Nombre del campo de fecha a usar, o None si no hay fecha de análisis disponible
    """
    tipo_analisis_lower = tipo_analisis.lower()
    
    # Mapeo de tipos de análisis a campos de fecha disponibles
    campos_fecha_por_tipo = {
        'dosn': {
            'campos': ['dosn_fecha_analisis', 'dosn_fecha_inia', 'dosn_fecha_inase'],
            'default': 'dosn_fecha_analisis'
        },
        'pureza': {
            'campos': ['fecha_inia', 'fecha_inase'],
            'default': 'fecha_inia'
        },
        'germinacion': {
            'campos': ['fecha_germinacion'],
            'default': 'fecha_germinacion'
        },
        'pms': {
            'campos': ['fecha_creacion'],  # Fallback
            'default': 'fecha_creacion'
        },
        'sanitario': {
            'campos': ['fechacreacion'],  # Fallback
            'default': 'fechacreacion'
        },
        'tetrazolio': {
            'campos': ['fecha_creacion'],  # Fallback
            'default': 'fecha_creacion'
        },
        'pureza_pnotatum': {
            'campos': ['fecha_creacion'],  # Fallback
            'default': 'fecha_creacion'
        }
    }
    
    # Si no se conoce el tipo, retornar None
    if tipo_analisis_lower not in campos_fecha_por_tipo:
        return None
    
    tipo_info = campos_fecha_por_tipo[tipo_analisis_lower]
    
    # Si se especifica un campo específico
    if campo_fecha and campo_fecha.lower() != 'auto':
        campo_solicitado = campo_fecha.lower()
        # Verificar si el campo solicitado está disponible para este tipo
        if campo_solicitado in tipo_info['campos']:
            return campo_solicitado
        # Si no está disponible, usar el default
        logger.warning(f"Campo de fecha '{campo_fecha}' no disponible para '{tipo_analisis}', usando '{tipo_info['default']}'")
        return tipo_info['default']
    
    # Detección automática: usar el campo default
    return tipo_info['default']

# ================================
# MÓDULO: PARSING DE IDs DE ANÁLISIS
# ================================
def parsear_analisis_ids(analisis_ids_str: str) -> Dict[str, List[int]]:
    """
    Parsea un string de IDs de análisis a un diccionario.
    
    Formato soportado:
    - 'dosn:1,2,3;pureza:5,6' → {'dosn': [1,2,3], 'pureza': [5,6]}
    - '1,2,3' → {} (requiere tabla para detección automática, no soportado aún)
    
    Args:
        analisis_ids_str: String con IDs en formato 'tipo:id1,id2;tipo2:id3'
    
    Returns:
        Diccionario mapeando tipo de análisis a lista de IDs
    """
    resultado = {}
    
    if not analisis_ids_str or not analisis_ids_str.strip():
        return resultado
    
    # Dividir por punto y coma para obtener cada tipo
    partes = analisis_ids_str.split(';')
    
    for parte in partes:
        parte = parte.strip()
        if not parte:
            continue
        
        # Buscar dos puntos para separar tipo de IDs
        if ':' in parte:
            tipo, ids_str = parte.split(':', 1)
            tipo = tipo.strip().lower()
            ids_str = ids_str.strip()
            
            if tipo and ids_str:
                # Parsear IDs como enteros, filtrando los inválidos
                ids = []
                for id_str in ids_str.split(','):
                    id_str = id_str.strip()
                    if id_str:
                        try:
                            ids.append(int(id_str))
                        except ValueError:
                            logger.warning(f"ID inválido ignorado: '{id_str}' para tipo '{tipo}'")
                            continue
                
                if ids:
                    resultado[tipo] = ids
        else:
            # Formato simple sin tipo (no soportado aún, requiere tabla)
            logger.warning(f"Formato de IDs sin tipo no soportado: '{parte}'. Use formato 'tipo:id1,id2'")
    
    return resultado

# ================================
# MÓDULO: FUNCIONES AUXILIARES EXCEL
# ================================
def crear_workbook_excel(titulo: str) -> tuple:
    """Crea un nuevo Workbook de Excel. Retorna (wb, ws)."""
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl no está instalado")
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]  # límite de Excel
    return wb, ws

def obtener_estilo_encabezado():
    """Retorna el estilo para los encabezados."""
    if not OPENPYXL_AVAILABLE:
        return None
    
    # Color de fondo azul claro profesional
    fill = PatternFill(
        start_color="4472C4",  # Azul corporativo
        end_color="4472C4",
        fill_type="solid"
    )
    
    # Fuente en negrita, blanca, tamaño 11
    font = Font(
        bold=True,
        size=11,
        color="FFFFFF"  # Blanco
    )
    
    # Bordes medianos negros
    border = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )
    
    # Alineación centrada
    alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    
    return {
        'fill': fill,
        'font': font,
        'border': border,
        'alignment': alignment
    }

def obtener_estilo_datos():
    """Retorna el estilo para las celdas de datos."""
    if not OPENPYXL_AVAILABLE:
        return None
    
    # Bordes delgados negros
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Alineación vertical centrada
    alignment = Alignment(
        vertical='center',
        wrap_text=True
    )
    
    # Fuente normal
    font = Font(size=11)
    
    return {
        'border': border,
        'alignment': alignment,
        'font': font
    }

def escribir_encabezados_excel(ws, encabezados: list):
    """Escribe los encabezados en una hoja de Excel con formato profesional."""
    ws.append(encabezados)
    
    if not OPENPYXL_AVAILABLE:
        return
    
    estilo = obtener_estilo_encabezado()
    if estilo:
        # Aplicar estilo a cada celda del encabezado
        for col_idx, header in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = estilo['fill']
            cell.font = estilo['font']
            cell.border = estilo['border']
            cell.alignment = estilo['alignment']
        
        # Ajustar altura de la fila de encabezado
        ws.row_dimensions[1].height = 25

def escribir_filas_excel(ws, rows: list):
    """Escribe las filas de datos en una hoja de Excel con formato profesional."""
    # #region agent log
    import json as _json_log
    try:
        with open(r"c:\Github\IniaProject\.cursor\debug.log", "a", encoding="utf-8") as _f:
            _f.write(_json_log.dumps({"location": "ExportExcel.py:escribir_filas_excel", "message": "Starting to write rows", "data": {"row_count": len(rows)}, "timestamp": __import__('time').time(), "sessionId": "debug-session", "hypothesisId": "D"}) + "\n")
    except: pass
    # #endregion
    if not OPENPYXL_AVAILABLE:
        for row in rows:
            values = [serialize_value(value) for value in row]
            ws.append(values)
        return
    
    estilo = obtener_estilo_datos()
    start_row = ws.max_row + 1
    
    for row_num, row in enumerate(rows, start=0):
        # Guardar valores originales para determinar tipo
        original_values = list(row)
        values = [serialize_value(value) for value in row]
        # #region agent log
        try:
            with open(r"c:\Github\IniaProject\.cursor\debug.log", "a", encoding="utf-8") as _f:
                _f.write(_json_log.dumps({"location": "ExportExcel.py:escribir_filas_excel_before_append", "message": f"Appending row {row_num}", "data": {"row_num": row_num}, "timestamp": __import__('time').time(), "sessionId": "debug-session", "hypothesisId": "D"}) + "\n")
        except: pass
        # #endregion
        try:
            ws.append(values)
        except Exception as _append_err:
            # #region agent log
            try:
                with open(r"c:\Github\IniaProject\.cursor\debug.log", "a", encoding="utf-8") as _f:
                    _f.write(_json_log.dumps({"location": "ExportExcel.py:escribir_filas_excel_append_error", "message": "ERROR appending row", "data": {"row_num": row_num, "error": str(_append_err), "values_repr": [repr(v)[:100] for v in values]}, "timestamp": __import__('time').time(), "sessionId": "debug-session", "hypothesisId": "D"}) + "\n")
            except: pass
            # #endregion
            raise
        row_idx = start_row + row_num
        
        if estilo:
            # Aplicar estilo a cada celda de la fila
            for col_idx, (original_value, serialized_value) in enumerate(zip(original_values, values), start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = estilo['border']
                cell.font = estilo['font']
                
                # Alineación horizontal según el tipo de dato original
                if isinstance(original_value, (int, float)):
                    cell.alignment = Alignment(
                        horizontal='right',
                        vertical='center',
                        wrap_text=True
                    )
                elif isinstance(original_value, (datetime, date)):
                    cell.alignment = Alignment(
                        horizontal='center',
                        vertical='center',
                        wrap_text=True
                    )
                else:
                    cell.alignment = Alignment(
                        horizontal='left',
                        vertical='center',
                        wrap_text=True
                    )
        
        # Altura de fila estándar
        ws.row_dimensions[row_idx].height = 18

def ajustar_ancho_columnas_excel(ws, columnas: list, min_width: int = 12, max_width: int = 50):
    """Ajusta el ancho de las columnas en una hoja de Excel."""
    for idx, col_name in enumerate(columnas, start=1):
        # Calcular el ancho basado en el contenido
        max_len = len(str(col_name))  # Empezar con el ancho del encabezado
        
        # Revisar todas las filas de datos
        for r in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=r, column=idx).value
            if cell_value is not None:
                cell_len = len(str(cell_value))
                max_len = max(max_len, cell_len)
        
        # Ajustar ancho con un poco de padding
        calculated_width = min(max(min_width, max_len + 3), max_width)
        ws.column_dimensions[get_column_letter(idx)].width = calculated_width

def guardar_workbook_excel(wb, xlsx_path: str) -> bool:
    """Guarda un Workbook de Excel en un archivo."""
    try:
        wb.save(xlsx_path)
        return True
    except Exception as e:
        log_fail(f"Error guardando Excel: {e}")
        return False

# ================================
# MÓDULO: EXPORTACIÓN EXCEL GENÉRICA
# ================================
def export_analisis_generico(session, model, xlsx_path: str, filtro_where: Optional[str] = None, filtro_params: Optional[Dict[str, Any]] = None) -> str:
    """
    Exporta otros análisis con formato genérico.
    
    Args:
        session: Sesión de SQLAlchemy
        model: Modelo de la tabla
        xlsx_path: Ruta del archivo Excel a generar
        filtro_where: Cláusula WHERE adicional (opcional)
        filtro_params: Parámetros para la cláusula WHERE (opcional)
    
    Returns:
        Ruta del archivo generado o cadena vacía si falla
    """
    try:
        if not OPENPYXL_AVAILABLE:
            log_fail("openpyxl no está instalado")
            return ""
        
        # Obtener y validar columnas
        columnas_validas, columnas_analisis = obtener_columnas_validas(session, model)
        if not columnas_analisis:
            return ""
        
        # Obtener nombre de tabla
        tabla_nombre = obtener_nombre_tabla(model)
        
        # Obtener datos con filtros opcionales
        rows = obtener_datos_tabla(session, tabla_nombre, columnas_analisis, filtro_where, filtro_params)
        
        # Crear workbook
        wb, ws = crear_workbook_excel(tabla_nombre)
        
        # Escribir datos
        escribir_encabezados_excel(ws, columnas_analisis)
        escribir_filas_excel(ws, rows)
        
        # Ajustar columnas
        ajustar_ancho_columnas_excel(ws, columnas_analisis)
        
        # Guardar
        if guardar_workbook_excel(wb, xlsx_path):
            log_ok(f"Archivo generado: {xlsx_path} ({len(rows)} filas)")
            return xlsx_path
        return ""
    except Exception as e:
        tabla_nombre = obtener_nombre_tabla(model)
        log_fail(f"No se pudo exportar {tabla_nombre} a Excel: {e}")
        return ""

def export_table_xlsx(session, model, output_dir: str) -> str:
    """Exporta una tabla a formato Excel."""
    table = obtener_nombre_tabla(model)
    # Normalizar nombre de archivo: siempre lowercase para evitar problemas con alias
    table_normalized = table.lower()
    xlsx_path = os.path.join(output_dir, f"{table_normalized}.xlsx")
    log_step(f"➡️ Exportando {table} a Excel...")
    
    try:
        if not OPENPYXL_AVAILABLE:
            log_fail("openpyxl no está instalado")
            return ""
        
        # Usar formato genérico para todos los análisis
        return export_analisis_generico(session, model, xlsx_path)
    except Exception as e:
        log_fail(f"No se pudo exportar {table} a Excel: {e}")
        return ""

# ================================
# MÓDULO: EXPORTACIÓN CON FILTROS
# ================================
def export_analisis_filtrados(
    session,
    tipos_analisis: List[str],
    analisis_ids: Optional[Dict[str, List[int]]] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    campo_fecha: Optional[str] = None,
    output_dir: str = "exports",
    fmt: str = "xlsx"
) -> List[str]:
    """
    Exporta análisis con filtros avanzados (IDs y fechas).
    
    Args:
        session: Sesión de SQLAlchemy
        tipos_analisis: Lista de tipos de análisis a exportar (ej: ['dosn', 'pureza'])
        analisis_ids: Diccionario mapeando tipo de análisis a lista de IDs
                     (ej: {'dosn': [1, 2, 3], 'pureza': [5, 6]})
        fecha_desde: Fecha de inicio del rango (opcional)
        fecha_hasta: Fecha de fin del rango (opcional)
        campo_fecha: Campo de fecha específico a usar ('auto' para detección automática)
        output_dir: Directorio de salida
        fmt: Formato de exportación ('xlsx' o 'csv')
    
    Returns:
        Lista de rutas de archivos generados
    """
    archivos_generados = []
    
    try:
        if not OPENPYXL_AVAILABLE:
            log_fail("openpyxl no está instalado")
            return archivos_generados
        
        # Mapeo de nombres de tipos a nombres de tablas
        tipo_a_tabla = {
            'dosn': 'dosn',
            'pureza': 'pureza',
            'germinacion': 'germinacion',
            'pms': 'pms',
            'sanitario': 'sanitario',
            'tetrazolio': 'tetrazolio',
            'pureza_pnotatum': 'pureza_pnotatum'
        }
        
        # Mapeo de tipos a nombres de columna ID
        tipo_a_id_col = {
            'dosn': 'dosn_id',
            'pureza': 'pureza_id',
            'germinacion': 'germinacion_id',
            'pms': 'pms_id',
            'sanitario': 'sanitario_id',
            'tetrazolio': 'tetrazolio_id',
            'pureza_pnotatum': 'pureza_pnotatum_id'
        }
        
        # Mapeo de tipos a nombres de columna active
        tipo_a_activo_col = {
            'dosn': 'dosn_activo',
            'pureza': 'pureza_activo',
            'germinacion': 'germinacion_activo',
            'pms': 'pms_activo',
            'sanitario': 'sanitario_activo',
            'tetrazolio': 'tetrazolio_activo',
            'pureza_pnotatum': 'pureza_activo'
        }
        
        for tipo_analisis in tipos_analisis:
            tipo_lower = tipo_analisis.lower()
            
            # Obtener nombre de tabla
            tabla_nombre = tipo_a_tabla.get(tipo_lower)
            if not tabla_nombre:
                log_fail(f"Tipo de análisis desconocido: {tipo_analisis}")
                continue
            
            # Intentar obtener el modelo
            try:
                model = obtener_modelo(tabla_nombre)
            except (AttributeError, KeyError) as e:
                log_fail(f"No se pudo obtener modelo para {tabla_nombre}: {e}")
                continue
            
            # Construir filtros WHERE
            condiciones_where = []
            params = {}
            
            # Filtro por IDs si se proporciona
            if analisis_ids and tipo_lower in analisis_ids:
                ids = analisis_ids[tipo_lower]
                if ids:
                    id_col = tipo_a_id_col.get(tipo_lower, f"{tipo_lower}_id")
                    # Verificar que la columna existe
                    columnas_validas, _ = obtener_columnas_validas(session, model)
                    if id_col in columnas_validas:
                        placeholders = ','.join([f':id_{i}' for i in range(len(ids))])
                        condiciones_where.append(f"{id_col} IN ({placeholders})")
                        for i, id_val in enumerate(ids):
                            params[f'id_{i}'] = id_val
            
            # Filtro por fechas si se proporciona
            campo_fecha_analisis = obtener_campo_fecha_analisis(tipo_lower, campo_fecha)
            if campo_fecha_analisis and (fecha_desde or fecha_hasta):
                # Verificar que el campo existe
                columnas_validas, _ = obtener_columnas_validas(session, model)
                if campo_fecha_analisis in columnas_validas:
                    if fecha_desde:
                        condiciones_where.append(f"{campo_fecha_analisis} >= :fecha_desde")
                        params['fecha_desde'] = fecha_desde
                    if fecha_hasta:
                        condiciones_where.append(f"{campo_fecha_analisis} <= :fecha_hasta")
                        params['fecha_hasta'] = fecha_hasta
            
            # Filtro por active = true (solo exportar análisis activos)
            activo_col = tipo_a_activo_col.get(tipo_lower)
            if activo_col:
                # Verificar que la columna existe
                columnas_validas, _ = obtener_columnas_validas(session, model)
                if activo_col in columnas_validas:
                    condiciones_where.append(f"{activo_col} = true")
            
            # Construir cláusula WHERE completa
            filtro_where = None
            filtro_params = None
            if condiciones_where:
                filtro_where = " AND ".join(condiciones_where)
                filtro_params = params
            
            # Generar nombre de archivo
            tabla_normalized = tabla_nombre.lower()
            xlsx_path = os.path.join(output_dir, f"{tabla_normalized}.xlsx")
            
            # Exportar con filtros
            log_step(f"Exportando {tabla_nombre} con filtros...")
            archivo_generado = export_analisis_generico(
                session, 
                model, 
                xlsx_path, 
                filtro_where=filtro_where,
                filtro_params=filtro_params
            )
            
            if archivo_generado:
                archivos_generados.append(archivo_generado)
        
        return archivos_generados
        
    except Exception as e:
        log_fail(f"Error en export_analisis_filtrados: {e}")
        return archivos_generados

def export_analisis_por_lote(
    session,
    lote_id: int,
    output_dir: str = "exports",
    fmt: str = "xlsx"
) -> List[str]:
    """
    Exporta todos los análisis asociados a un lote específico.
    
    Args:
        session: Sesión de SQLAlchemy
        lote_id: ID del lote del cual exportar análisis
        output_dir: Directorio de salida
        fmt: Formato de exportación ('xlsx' o 'csv')
    
    Returns:
        Lista de rutas de archivos generados
    """
    archivos_generados = []
    
    try:
        if not OPENPYXL_AVAILABLE:
            log_fail("openpyxl no está instalado")
            return archivos_generados
        
        # Obtener recibo_id asociado al lote_id
        query_recibo = text("""
            SELECT RECIBO_ID 
            FROM RECIBO 
            WHERE LOTE_ID = :lote_id 
            AND RECIBO_ACTIVO = true
            LIMIT 1
        """)
        result_recibo = session.execute(query_recibo, {"lote_id": lote_id}).fetchone()
        
        if not result_recibo or not result_recibo[0]:
            log_fail(f"No se encontró recibo activo asociado al lote {lote_id}")
            return archivos_generados
        
        recibo_id = result_recibo[0]
        log_step(f"Recibo encontrado para lote {lote_id}: RECIBO_ID = {recibo_id}")
        
        # Mapeo de tipos de análisis a nombres de tablas y columnas
        tipos_analisis = {
            'dosn': {'tabla': 'dosn', 'id_col': 'dosn_id', 'recibo_col': 'recibo_id', 'activo_col': 'dosn_activo'},
            'pureza': {'tabla': 'pureza', 'id_col': 'pureza_id', 'recibo_col': 'recibo_id', 'activo_col': 'pureza_activo'},
            'germinacion': {'tabla': 'germinacion', 'id_col': 'germinacion_id', 'recibo_col': 'recibo_id', 'activo_col': 'germinacion_activo'},
            'pms': {'tabla': 'pms', 'id_col': 'pms_id', 'recibo_col': 'recibo_id', 'activo_col': 'pms_activo'},
            'sanitario': {'tabla': 'sanitario', 'id_col': 'sanitario_id', 'recibo_col': 'sanitario_reciboid', 'activo_col': 'sanitario_activo'},
            'tetrazolio': {'tabla': 'tetrazolio', 'id_col': 'tetrazolio_id', 'recibo_col': 'recibo_id', 'activo_col': 'tetrazolio_activo'},
            'pureza_pnotatum': {'tabla': 'pureza_pnotatum', 'id_col': 'pureza_pnotatum_id', 'recibo_col': 'recibo_id', 'activo_col': 'pureza_activo'}
        }
        
        # Exportar cada tipo de análisis asociado al recibo
        for tipo, config in tipos_analisis.items():
            tabla_nombre = config['tabla']
            recibo_col = config['recibo_col']
            
            # Verificar si la tabla existe y tiene datos para este recibo
            try:
                model = obtener_modelo(tabla_nombre)
            except (AttributeError, KeyError) as e:
                log_step(f"Tabla {tabla_nombre} no encontrada o sin modelo, omitiendo...")
                continue
            
            # Verificar que existe la columna recibo_id
            columnas_validas, _ = obtener_columnas_validas(session, model)
            if recibo_col not in columnas_validas:
                log_step(f"Tabla {tabla_nombre} no tiene columna {recibo_col}, omitiendo...")
                continue
            
            # Construir filtro WHERE para filtrar por recibo_id y active = true
            condiciones_where = [f"{recibo_col} = :recibo_id"]
            filtro_params = {'recibo_id': recibo_id}
            
            # Agregar filtro de active = true si la columna existe
            activo_col = config.get('activo_col')
            if activo_col and activo_col in columnas_validas:
                condiciones_where.append(f"{activo_col} = true")
            
            filtro_where = " AND ".join(condiciones_where)
            
            # Verificar si hay datos antes de exportar
            query_count = text(f"SELECT COUNT(*) FROM {tabla_nombre} WHERE {filtro_where}")
            count_result = session.execute(query_count, filtro_params).fetchone()
            if not count_result or count_result[0] == 0:
                log_step(f"No hay análisis de tipo {tipo} para el recibo {recibo_id}, omitiendo...")
                continue
            
            # Generar nombre de archivo
            tabla_normalized = tabla_nombre.lower()
            xlsx_path = os.path.join(output_dir, f"{tabla_normalized}.xlsx")
            
            # Exportar con filtro de recibo
            log_step(f"Exportando {tabla_nombre} para lote {lote_id} (recibo {recibo_id})...")
            archivo_generado = export_analisis_generico(
                session, 
                model, 
                xlsx_path, 
                filtro_where=filtro_where,
                filtro_params=filtro_params
            )
            
            if archivo_generado:
                archivos_generados.append(archivo_generado)
        
        if archivos_generados:
            log_ok(f"Exportación completada para lote {lote_id}: {len(archivos_generados)} archivo(s) generado(s)")
        else:
            log_fail(f"No se generaron archivos para el lote {lote_id}")
        
        return archivos_generados
        
    except Exception as e:
        log_fail(f"Error en export_analisis_por_lote: {e}")
        return archivos_generados

# ================================
# MÓDULO: EXPORTACIÓN PRINCIPAL
# ================================
def export_selected_tables(tables: list, output_dir: str, fmt: str, incluir_sin_pk: bool = True) -> None:
    """Exporta las tablas seleccionadas a Excel.
    
    Args:
        tables: Lista de nombres de tablas a exportar. Si está vacía, exporta todas.
        output_dir: Directorio de salida para los archivos
        fmt: Formato de exportación ('xlsx' o 'csv')
        incluir_sin_pk: Si es True, incluye también tablas sin Primary Key
    """
    try:
        engine = obtener_engine()
        
        # Inicializar automapeo antes de crear la sesión
        log_step("Inicializando automapeo de la base de datos...")
        inicializar_automap(engine)
        log_step(f"Modelos disponibles: {list(MODELS.keys())}")
        
        Session = sessionmaker(bind=engine)
        session = Session()
        
        # Probar la conexión
        session.execute(text("SELECT 1"))
        
        try:
            out_dir = ensure_output_dir(output_dir)
            exported = 0
            tablas_a_exportar = set()
            
            # Si no se especifican tablas, usar todas las disponibles
            if not tables:
                tablas_a_exportar = set(MODELS.keys())
            else:
                tablas_a_exportar = set(t.lower() for t in tables)
            
            # Si se debe incluir tablas sin PK, agregarlas a la lista
            if incluir_sin_pk:
                # Primero, agregar tablas relacionadas de las tablas principales que se están exportando
                tablas_principales = list(tablas_a_exportar.copy())
                for tabla_principal in tablas_principales:
                    try:
                        tablas_relacionadas = obtener_tablas_relacionadas(session, tabla_principal)
                        for tabla_rel in tablas_relacionadas:
                            tabla_rel_lower = tabla_rel.lower()
                            # Excluir certificado explícitamente
                            if tabla_rel_lower != 'certificado' and tabla_rel_lower not in tablas_a_exportar:
                                log_step(f"Incluyendo tabla relacionada con {tabla_principal}: {tabla_rel}")
                                tablas_a_exportar.add(tabla_rel_lower)
                    except Exception as e:
                        log_fail(f"Error obteniendo tablas relacionadas con {tabla_principal}: {e}")
                        # Hacer rollback para continuar
                        try:
                            session.rollback()
                        except:
                            pass
                        continue
                
                # Luego, agregar todas las demás tablas sin PK
                try:
                    tablas_sin_pk = obtener_tablas_sin_pk(session)
                    for tabla_sin_pk in tablas_sin_pk:
                        tabla_lower = tabla_sin_pk.lower()
                        # Excluir certificado explícitamente
                        if tabla_lower != 'certificado' and tabla_lower not in tablas_a_exportar:
                            # Intentar obtener el modelo o crear uno dinámico
                            if tabla_lower in MODELS:
                                tablas_a_exportar.add(tabla_lower)
                            else:
                                # Si no está en MODELS, intentar agregarla directamente
                                log_step(f"Incluyendo tabla sin PK: {tabla_sin_pk}")
                                tablas_a_exportar.add(tabla_lower)
                except Exception as e:
                    log_fail(f"Error obteniendo tablas sin PK: {e}")
                    # Hacer rollback para continuar
                    try:
                        session.rollback()
                    except:
                        pass
            
            # Excluir certificado explícitamente de la lista final
            tablas_a_exportar = {t for t in tablas_a_exportar if t.lower() != 'certificado'}
            
            log_step(f"Total de tablas a exportar: {len(tablas_a_exportar)}")
            
            for name in tablas_a_exportar:
                try:
                    # Intentar obtener el modelo
                    model = None
                    try:
                        model = obtener_modelo(name)
                    except (AttributeError, KeyError):
                        # Si no se encuentra el modelo, intentar exportar directamente
                        log_step(f"Tabla {name} no encontrada en modelos, intentando exportación directa...")
                        # Verificar que la tabla existe en la BD
                        query_check = text("""
                            SELECT EXISTS (
                                SELECT 1
                                FROM information_schema.tables
                                WHERE table_schema = 'public'
                                    AND table_name = :tabla
                            )
                        """)
                        exists = session.execute(query_check, {"tabla": name}).fetchone()[0]
                        if not exists:
                            log_fail(f"Tabla {name} no existe en la base de datos")
                            continue
                        
                        # Crear un modelo temporal para la exportación
                        # Usar el modelo genérico si es posible
                        if name in MODELS:
                            model = MODELS[name]
                        else:
                            # Exportar directamente sin modelo
                            log_step(f"Exportando tabla {name} directamente (sin modelo)...")
                            try:
                                # Obtener columnas de la tabla
                                columnas = verificar_estructura_tabla(session, name)
                                if not columnas:
                                    log_fail(f"No se pudieron obtener columnas de {name}")
                                    continue
                                
                                # Filtrar columnas de análisis
                                columnas_analisis = filtrar_columnas_analisis(columnas, name)
                                if not columnas_analisis:
                                    log_step(f"Tabla {name} no tiene columnas de análisis después del filtrado")
                                    continue
                                
                                # Obtener datos
                                rows = obtener_datos_tabla(session, name, columnas_analisis)
                                
                                # Crear workbook
                                if not OPENPYXL_AVAILABLE:
                                    log_fail("openpyxl no está instalado")
                                    continue
                                
                                wb, ws = crear_workbook_excel(name)
                                
                                # Escribir datos
                                escribir_encabezados_excel(ws, columnas_analisis)
                                escribir_filas_excel(ws, rows)
                                
                                # Ajustar columnas
                                ajustar_ancho_columnas_excel(ws, columnas_analisis)
                                
                                # Guardar (normalizar nombre a lowercase)
                                name_normalized = name.lower()
                                xlsx_path = os.path.join(out_dir, f"{name_normalized}.xlsx")
                                if guardar_workbook_excel(wb, xlsx_path):
                                    log_ok(f"Archivo generado: {xlsx_path} ({len(rows)} filas)")
                                    exported += 1
                                continue
                            except Exception as e:
                                log_fail(f"Error exportando tabla {name} directamente: {e}")
                                continue
                    
                    if not model:
                        log_fail(f"No se pudo obtener modelo para tabla: {name}")
                        continue
                    
                    if fmt == "xlsx":
                        path = export_table_xlsx(session, model, out_dir)
                    else:
                        log_fail(f"Formato {fmt} no soportado. Solo se soporta xlsx.")
                        continue
                    
                    if path:
                        exported += 1
                except Exception as e:
                    log_fail(f"Error exportando tabla {name}: {e}")
                    continue
            
            log_ok(f"Tablas exportadas correctamente: {exported}/{len(tablas_a_exportar)}")
            
            if exported == 0:
                raise Exception("No se pudo exportar ninguna tabla")
                
        finally:
            session.close()
    except Exception as e:
        log_fail(f"Error en export_selected_tables: {e}")
        raise

# ================================
# MÓDULO: FUNCIÓN PRINCIPAL
# ================================
def main():
    """Función principal del script."""
    parser = argparse.ArgumentParser(description="Exporta análisis del proyecto INIA a Excel")
    parser.add_argument(
        "--tables",
        nargs="*",
        default=[],
        help="Lista de análisis a exportar. Por defecto exporta todos los análisis disponibles"
    )
    parser.add_argument(
        "--out",
        default=os.path.join(os.path.dirname(__file__), "exports"),
        help="Directorio de salida para los archivos"
    )
    parser.add_argument(
        "--format",
        choices=["xlsx"],
        default="xlsx",
        help="Formato de salida (xlsx por defecto)"
    )
    args = parser.parse_args()

    # Determinar si incluir tablas sin PK
    incluir_sin_pk = True  # Por defecto incluir tablas sin PK
    
    # Verificar si hay argumentos para excluir
    if hasattr(args, 'excluir_sin_pk') and args.excluir_sin_pk:
        incluir_sin_pk = False
    elif hasattr(args, 'incluir_sin_pk') and not args.incluir_sin_pk:
        incluir_sin_pk = False

    log_step(f"Iniciando exportación de análisis en formato {args.format}…")
    log_step("Se exportarán datos de análisis (excluyendo IDs, estados activos, fechas de control)")
    if incluir_sin_pk:
        log_step("Incluyendo tablas sin Primary Key (tablas vinculadas)")
    export_selected_tables(args.tables, args.out, args.format, incluir_sin_pk=incluir_sin_pk)

if __name__ == "__main__":
    logging.getLogger().setLevel(logging.INFO)
    main()

# ================================
# RE-EXPORTS PARA COMPATIBILIDAD
# ================================
# Mantener re-exports para que código existente que importa desde este módulo siga funcionando
from db_common import Base, MODELS, obtener_engine, inicializar_automap, obtener_modelo, obtener_nombre_tabla_seguro
